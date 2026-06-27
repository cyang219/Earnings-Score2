import io
import json
import re
import time

import openpyxl
import streamlit as st
from anthropic import Anthropic
from openpyxl.utils import column_index_from_string

MODEL = "claude-sonnet-4-6"
THEMES_EFFORT = "medium"
READTHROUGH_MODEL = "claude-sonnet-4-6"
READTHROUGH_EFFORT = "medium"
THEME_DELTA_MODEL = "claude-sonnet-4-6"
THEME_DELTA_EFFORT = "low"

MIN_QUARTERS = 2

COMMON_THEME_SET = [
    "Top-line Growth (e.g. Revenue, Bookings, GMV, ARR, etc.)",
    "Bottom-line Expansion/Contraction (e.g. Margins, Return on Investment, Take-rate, etc.)",
    "Financial Guidance and Forward-looking Business Commentary",
]


# ============================================================
# PROMPTS
# ============================================================

def build_quarter_labels(n: int) -> list[str]:
    return [f"Q-{n - 1 - i}" for i in range(n - 1)] + ["Latest_Q"]


def _oxford_join(items: list[str]) -> str:
    if len(items) == 1:
        return items[0]
    if len(items) == 2:
        return f"{items[0]} and {items[1]}"
    return ", ".join(items[:-1]) + f", and {items[-1]}"


def build_theme_system_prompt(labels: list[str], existing_themes: list[str] = None) -> str:
    n = len(labels)
    annotated_labels = [f"{labels[0]} (oldest)"] + labels[1:-1] + [f"{labels[-1]} (most recent)"]
    label_description = _oxford_join(annotated_labels)
    delta_list = ", ".join(f"{labels[i]} vs {labels[i + 1]}" for i in range(n - 1))
    json_skeleton = "{\n" + ",\n".join(
        f'  "<period of {label}>": {{"themes": ["...", "...", "..."]}}' for label in labels[1:]
    ) + "\n}"

    step5 = ""
    existing_section = ""
    if existing_themes:
        step5 = "\n5. HISTORICAL THEME RECONCILIATION. After completing step 4, compare each selected theme name against the EXISTING THEMES list at the end of this prompt. If a selected theme is semantically the same underlying topic as an existing theme, rename it to match that existing name exactly (preserve its exact spelling and capitalization). Apply this only when the match is strong — the two names clearly describe the same investment topic. If no existing theme is a close semantic match, keep the name from step 4. Do not force a weak or partial match."
        existing_section = "\n\nEXISTING THEMES FOR RECONCILIATION (step 5):\n" + "\n".join(f"- {t}" for t in existing_themes)

    return f"""You are given {n} quarters of earnings call Q&A bullet-point summaries for the same company, ordered chronologically and labeled {label_description}.

For each consecutive delta ({delta_list}):
1. Find 3 common themes that are heavily discussed during the Q&A sessions of both quarters in that delta.
2. Find 2 new or emerging themes that are heavily discussed during the Q&A session of the later quarter in that delta but were lightly or never discussed during the earlier quarter.
3. Rank all 5 of those themes (the 3 common plus the 2 emerging) by how much discussion volume each received in the LATER quarter's Q&A session only (ignore the earlier quarter's volume for this ranking). Keep only the top 3 themes by that ranking, ordered from highest to lowest volume.
4. CROSS-DELTA CONSOLIDATION. After the top-3 themes have been selected for every delta, compare theme names across deltas (not within the same delta — that's already handled below). If a theme chosen for one delta is semantically the same underlying topic as a theme chosen for another delta, rename both occurrences to one common, consistent name (2-3 words, high-level and neutral, per the Rules below) so the same underlying theme is tracked identically wherever it recurs. Do not change the ranking, the order, or which themes were selected — only normalize the name.{step5}

Rules:
- "Heavily discussed" / theme volume is judged by the amount of Q&A dialogue (number of words) spent discussing the theme.
- Each theme must be very high-level and neutral (e.g. "Proprietary Silicon", not "Proprietary Silicon Development Timing").
- Each theme in a delta quarter must be over 80% different semantically compared to the others.
- Each theme name must be 2-3 words.

Output format - return ONLY valid JSON, no commentary, no markdown code fences, no preamble. Structure:
{json_skeleton}
Each delta's key in the JSON must be the period label (e.g. "Q1 2026") of the LATER quarter in that delta. Each delta's "themes" list must contain exactly 3 entries: the top 3 themes after the step-3 ranking and cut, ordered from highest to lowest discussion volume in the later quarter, with no indication of which were originally common vs. emerging.{existing_section}"""


READTHROUGH_DESCRIPTION_LENGTH_RULE = (
    "The two fields combined must be 350-400 characters long, including spaces and punctuation"
)

THEME_DELTA_PROMPT_TEMPLATE = """You are an earnings-call analyst comparing management messaging and analyst tone across investment themes for one quarter-over-quarter delta. You will be given consecutive telegraphic summaries of quarterly earnings calls from a single publicly traded company.

DEFINITIONS (canonical — all later references point back here):
- MANAGEMENT MESSAGE = what management conveys about a theme in the prepared remarks and their Q&A answers: the substance, confidence, and framing of their commentary on that theme.
- ANALYST TONE = the stance reflected in analysts' Q&A questions on a theme: whether they probe risk, press skeptically, or signal confidence, and how management's answers are received.
- DELTA DIRECTION = the FIRST-named quarter in QUARTERS TO ANALYZE is the LATER quarter; the SECOND-named is the EARLIER quarter. A signal describes how the later quarter changed versus the earlier quarter.
- THEME INDEPENDENCE = each theme is analyzed in complete isolation. Evidence, tone, framing, or sentiment from one theme MUST NOT be transferred to another. Management and analyst sides are also judged independently per theme; they need not move in the same direction.
- DATA BOUNDARY = use ONLY the two quarters' transcript data provided. Do NOT use outside information.

THEME LIST:
{theme_list}

QUARTERS TO ANALYZE:
{quarter_delta}

TASK:

Step 1: PER-THEME READ. For each theme in THEME LIST, using only the two quarters (see DATA BOUNDARY), internally summarize the later quarter's management message and analyst tone for that theme, and how each compares to the earlier quarter. If a theme is absent in the earlier quarter, note this now so the special-case rule in Step 2 applies correctly. Process each theme fully and in isolation before moving to the next (see THEME INDEPENDENCE). Use this only as internal reasoning; do not output Step 1.

Step 2: SIGNAL SCRATCHPAD. Start a <scratchpad> assigning a management signal and an analyst signal to each theme. Judge each signal only by the change versus the earlier quarter (see DELTA DIRECTION), not by absolute positivity or negativity.

Classify each side (management, then analyst) for each theme using this rule, then assign the signal:
- Improved: a clear positive change vs the earlier quarter — management escalates confidence/affirms progress/resolves a prior concern on that theme, OR analyst questions shift from probing a risk toward confirming an improvement.
- Worsened: a clear negative change — management softens, walks back, or introduces a new concern on that theme, OR analyst questions surface a new risk management cannot fully address.
- Stable: no clear directional change — framing, confidence, and tone on that theme are materially unchanged, OR positive and negative shifts roughly offset.
An Improved or Worsened signal requires a specific, citable QoQ change on that theme for that side (see DATA BOUNDARY). If no such concrete change can be named, the signal must be Stable. When evidence is genuinely balanced or ambiguous, default to Stable rather than guessing a direction.
Special case — theme absent in the earlier quarter: if the theme is not discussed in the earlier quarter, judge by the later quarter's absolute message/tone: positive = Improved, negative = Worsened, neutral or limited = Stable. Do not mark Improved or Worsened solely because the theme appeared.

Copy the exact template below and fill in the brackets, producing one line per theme in THEME LIST order. Do NOT write anything else inside the scratchpad.
<scratchpad>
{scratchpad_format}
</scratchpad>
- Signal must be exactly one of: Improved, Worsened, Stable
- Each signal must be evidence-linked to that theme's own data (see THEME INDEPENDENCE)

Step 3: SIGNAL FIELD. Derive the "signal" field mechanically from the scratchpad Signal for that same theme and side (not an independent judgment):
- Improved -> "↑:"   |   Worsened -> "↓:"   |   Stable -> "→:"
If the field and the scratchpad Signal would ever disagree, the scratchpad Signal is authoritative.

Step 4: MESSAGE/TONE AND RATIONALE FIELDS. For each theme, produce three fields per side (mgmt and analyst): "signal" (the arrow from Step 3), then for mgmt "message" and for analyst "tone", and "rationale" for both:
- "message" (mgmt) / "tone" (analyst): one sentence stating the later-quarter management message or analyst tone on that theme.
- "rationale": one sentence (a) citing one specific QoQ change on that theme and (b) explaining why it improved, worsened, or stayed stable, matching that side's signal direction.
- {description_length_rule} across the "message"/"tone" field and its "rationale" field combined.
- STYLE: abridged, telegraphic style with abbreviations.
- Omit the final period "." at the end of every field.

Before producing the JSON, verify that (a) every theme in THEME LIST is present, (b) each theme's mgmt and analyst signals comply with THEME INDEPENDENCE, and (c) each rationale's described direction matches its signal arrow and the scratchpad Signal. If any disagree, the scratchpad Signal governs; revise to match.

JSON FORMAT: After </scratchpad>, output ONLY a valid JSON block using ```json tags, following the exact schema below, with one object per theme in THEME LIST order. Replace the theme placeholders with the exact theme names from THEME LIST. Do not output any text between </scratchpad> and the JSON block, or after the JSON block. Escape any double quotes inside field values.

```json
{json_format}
```"""


def build_delta_pairs(labels: list[str]) -> list[tuple[str, str]]:
    return [(labels[i + 1], labels[i]) for i in range(len(labels) - 1)]


def build_readthrough_quarter_mapping(ordered_quarters) -> str:
    lines = [f"{item['label']} = {item['period']}" for item in ordered_quarters]
    return "QUARTER MAPPING:\n" + "\n".join(lines)


MERGED_PROMPT_TEMPLATE = """You are an equity research analyst producing three signals per quarter-over-quarter delta from one company's consecutive earnings calls: a sector READ-THROUGH, a BULL forward case, and a BEAR forward case. Inputs are consecutive quarterly transcripts (or telegraphic summaries) ordered oldest to newest; the task is identical regardless of company, sector, fiscal calendar, or quarter count.

QUARTER MAPPING:
{quarter_mapping}
- DELTA: each adjacent pair is one delta; N quarters give N-1 deltas. In a "LATER_vs_EARLIER" key the first quarter is the later one, the second the earlier. Every signal describes how the later quarter changed versus the earlier.

DEFINITIONS:
- EXTERNAL BUSINESS FACTORS: industry/macro conditions management describes that read across to sector peers — end-market demand, customer spending, pricing, input/supply availability, competitive dynamics, regulatory/macro. EXCLUDES the company's own KPIs and internal initiatives (its revenue, margins, segments, roadmap, capital returns, restructuring). The READ-THROUGH uses external factors ONLY.
- FORWARD EXPECTATION: what a bull (or bear) investor would, on that quarter's call, expect for the business over the next 6–9 months. Infer from the whole call but prioritize the Q&A, where management is tested. Bull and bear are assessed separately and need not move in opposite directions.

RULES (apply to every delta and every signal):
- ANALYSIS INDEPENDENCE: read-through, bull, and bear are three independent reads of the same delta. Derive each from its own evidence; never soften, strengthen, or flip one because another track moved. They may point in different directions; sharing an underlying fact is fine as long as each is weighed on its own.
- DATA BOUNDARY: use ONLY the provided transcripts. No outside information.

TASK — for each delta, assign three signals (read-through, bull, bear), each Strengthened / Weakened / Stable:
- READ-THROUGH: select the single dominant external business factor. It MUST be discussed in both quarters of the delta — a factor absent from the earlier quarter is ineligible (no QoQ baseline to measure against). Among eligible factors, pick the one most relevant to sector peers, breaking ties by emphasis (airtime / strongest language). If no external factor appears in both quarters, the read-through is Stable. Then judge how the selected factor's read-through changed QoQ.
- BULL / BEAR: for each side, judge how that side's forward 6–9 month case changed QoQ, prioritizing Q&A evidence.

Signal rule (all three):
- Strengthened: a clear positive QoQ change — escalated language, a raised/affirmed external indicator or guidance, a broadened demand base, a resolved concern (bull/bear: analysts shift from probing a risk to confirming it, corroborated by management).
- Weakened: a clear negative QoQ change — softened/walked-back language, a new or amplified headwind, a narrowed demand base, a new risk management cannot fully address.
- Stable: no clear directional change, or offsetting shifts.
A Strengthened or Weakened signal requires a specific, citable QoQ change in direction — not a reaffirmation, an in-line/as-guided number, or a repeated stance. If none can be named, the signal is Stable. When evidence is ambiguous, default to Stable.
Map each signal to its arrow: Strengthened -> "↑:", Weakened -> "↓:", Stable -> "→:". The read-through arrow fills the "signal" field; each side's arrow fills its "topic" field.

OUTPUT FIELDS (per delta):
- READ-THROUGH "read_through": one sentence stating the later quarter's macro/industry read-through (external factors only). "rationale": one sentence citing one specific QoQ contrast and why it strengthened/weakened/held. {description_length_rule} across the two combined.
- BULL/BEAR "expectation" (per side): one sentence stating the later quarter's forward 6–9 month case. "context": one sentence citing one specific QoQ change in that side's case and why it strengthened/weakened/held. {description_length_rule} across the two combined.
- STYLE: abridged, telegraphic, abbreviations. Omit the final period "." at the end of every field.

Before producing JSON, verify for each delta that: (a) every rationale/context direction matches its arrow; (b) each ↑ or ↓ rests on a real directional QoQ change — else → (stable); (c) the read-through factor was discussed in both quarters; (d) no track's signal was adjusted to agree or disagree with another track's on the same delta.

JSON FORMAT: Output ONLY a valid JSON block using ```json tags, following the exact schema below, with one object per delta matching the periods and order in the schema. No text before or after the JSON block. Escape any double quotes inside field values.

```json
{json_format}
```"""


def build_merged_json_skeleton(labels: list[str]) -> str:
    entries = [
        f"""    {{
      "period": "{later}_vs_{earlier}",
      "signal": "↑: or ↓: or →:",
      "read_through": "one-sentence macro/industry read-through",
      "rationale": "one-sentence QoQ contrast explaining the read-through signal direction",
      "bull": {{
        "topic": "↑: or ↓: or →:",
        "expectation": "one-sentence bullish forward-looking expectation",
        "context": "one-sentence QoQ context explaining the bull signal direction"
      }},
      "bear": {{
        "topic": "↑: or ↓: or →:",
        "expectation": "one-sentence bearish forward-looking expectation",
        "context": "one-sentence QoQ context explaining the bear signal direction"
      }}
    }}"""
        for later, earlier in build_delta_pairs(labels)
    ]
    return "{\n  \"historical_analysis\": [\n" + ",\n".join(entries) + "\n  ]\n}"


def build_merged_system_prompt(ordered_quarters) -> str:
    labels = [item["label"] for item in ordered_quarters]
    return MERGED_PROMPT_TEMPLATE.format(
        quarter_mapping=build_readthrough_quarter_mapping(ordered_quarters),
        description_length_rule=READTHROUGH_DESCRIPTION_LENGTH_RULE,
        json_format=build_merged_json_skeleton(labels),
    )


def build_theme_list(themes: list[str]) -> str:
    return "\n".join(f"- {theme}" for theme in themes)


def build_quarter_delta_mapping(later_item, earlier_item) -> str:
    return (
        f"{later_item['label']} = {later_item['period']} (LATER)\n"
        f"{earlier_item['label']} = {earlier_item['period']} (EARLIER)"
    )


def build_theme_delta_scratchpad_format(themes: list[str]) -> str:
    return "\n".join(
        f"{theme} -> Mgmt Signal: [Signal]; Mgmt Driver: [2–5 word evidence]; "
        f"Analyst Signal: [Signal]; Analyst Driver: [2–5 word evidence]"
        for theme in themes
    )


def build_theme_delta_json_skeleton(themes: list[str]) -> str:
    entries = [
        f"""    {{
      "theme": "{theme}",
      "mgmt": {{
        "signal": "↑: or ↓: or →:",
        "message": "one-sentence management message",
        "rationale": "one-sentence QoQ contrast explaining the signal direction"
      }},
      "analyst": {{
        "signal": "↑: or ↓: or →:",
        "tone": "one-sentence analyst tone",
        "rationale": "one-sentence QoQ contrast explaining the signal direction"
      }}
    }}"""
        for theme in themes
    ]
    return "{\n  \"theme_analysis\": [\n" + ",\n".join(entries) + "\n  ]\n}"


def build_theme_delta_system_prompt(themes: list[str], later_item, earlier_item) -> str:
    return THEME_DELTA_PROMPT_TEMPLATE.format(
        theme_list=build_theme_list(themes),
        quarter_delta=build_quarter_delta_mapping(later_item, earlier_item),
        scratchpad_format=build_theme_delta_scratchpad_format(themes),
        description_length_rule=READTHROUGH_DESCRIPTION_LENGTH_RULE,
        json_format=build_theme_delta_json_skeleton(themes),
    )


# ============================================================
# ANALYSIS CALLS
# ============================================================

def extract_qa_section(content: str) -> str:
    match = re.search(r"QUESTION AND ANSWER SECTION", content, flags=re.IGNORECASE)
    return content[match.start():] if match else content


def build_themes_params(ordered_quarters, existing_themes: list[str] = None) -> dict:
    header = "Input Periods: " + ", ".join(
        f"{item['label']}:{item['period']}" for item in ordered_quarters
    )
    sections = "\n\n".join(
        f"=== {item['label']} ({item['period']}) ===\n{extract_qa_section(item['content'])}"
        for item in ordered_quarters
    )
    user_content = f"{header}\n\n{sections}"
    system_prompt = build_theme_system_prompt([item["label"] for item in ordered_quarters], existing_themes)
    return {
        "model": MODEL,
        "max_tokens": 64000,
        "thinking": {"type": "adaptive"},
        "output_config": {"effort": THEMES_EFFORT},
        "system": system_prompt,
        "messages": [{"role": "user", "content": user_content}],
    }


def parse_themes_response(raw_text: str):
    cleaned = re.sub(r"^```(?:json)?|```$", "", raw_text.strip(), flags=re.MULTILINE).strip()
    try:
        return json.loads(cleaned), None
    except json.JSONDecodeError as e:
        return None, f"Failed to parse JSON response: {e}\n\nRaw response:\n{raw_text}"


def parse_fenced_json_response(raw_text: str):
    match = re.search(r"```json\s*(.*?)```", raw_text, flags=re.DOTALL)
    if not match:
        return None, f"Failed to find JSON block in response\n\nRaw response:\n{raw_text}"
    try:
        return json.loads(match.group(1).strip()), None
    except json.JSONDecodeError as e:
        return None, f"Failed to parse JSON response: {e}\n\nRaw response:\n{raw_text}"


def build_merged_params(ordered_quarters) -> dict:
    sections = "\n\n".join(
        f"=== {item['label']} ({item['period']}) ===\n{item['content']}"
        for item in ordered_quarters
    )
    return {
        "model": READTHROUGH_MODEL,
        "max_tokens": 64000,
        "thinking": {"type": "adaptive"},
        "output_config": {"effort": READTHROUGH_EFFORT},
        "system": build_merged_system_prompt(ordered_quarters),
        "messages": [{"role": "user", "content": sections}],
    }


def build_theme_delta_params(themes: list[str], later_item, earlier_item) -> dict:
    sections = "\n\n".join(
        f"=== {item['label']} ({item['period']}) ===\n{item['content']}"
        for item in (later_item, earlier_item)
    )
    return {
        "model": THEME_DELTA_MODEL,
        "max_tokens": 64000,
        "thinking": {"type": "adaptive"},
        "output_config": {"effort": THEME_DELTA_EFFORT},
        "system": build_theme_delta_system_prompt(themes, later_item, earlier_item),
        "messages": [{"role": "user", "content": sections}],
    }


def build_themes_batch_request(custom_id: str, ordered_quarters, existing_themes: list[str] = None) -> dict:
    return {"custom_id": custom_id, "params": build_themes_params(ordered_quarters, existing_themes)}


def build_merged_batch_request(custom_id: str, ordered_quarters) -> dict:
    return {"custom_id": custom_id, "params": build_merged_params(ordered_quarters)}


def build_theme_delta_batch_request(custom_id: str, themes: list[str], later_item, earlier_item) -> dict:
    return {"custom_id": custom_id, "params": build_theme_delta_params(themes, later_item, earlier_item)}


def _stream_text(client: Anthropic, params: dict) -> str:
    with client.messages.stream(**params) as stream:
        response = stream.get_final_message()
    return "".join(block.text for block in response.content if block.type == "text")


def analyze_themes(client: Anthropic, ordered_quarters, existing_themes: list[str] = None):
    raw_text = _stream_text(client, build_themes_params(ordered_quarters, existing_themes))
    return parse_themes_response(raw_text)


def analyze_merged(client: Anthropic, ordered_quarters):
    raw_text = _stream_text(client, build_merged_params(ordered_quarters))
    return parse_fenced_json_response(raw_text)


def analyze_theme_delta(client: Anthropic, themes: list[str], later_item, earlier_item):
    raw_text = _stream_text(client, build_theme_delta_params(themes, later_item, earlier_item))
    return parse_fenced_json_response(raw_text)


# ============================================================
# BATCH PROCESSING
# ============================================================

def submit_stage1_batch(client: Anthropic, ordered_quarters, existing_themes: list[str] = None) -> str:
    requests = [
        build_themes_batch_request("themes", ordered_quarters, existing_themes),
        build_merged_batch_request("merged", ordered_quarters),
    ]
    batch = client.messages.batches.create(requests=requests)
    return batch.id


def submit_stage2_batch(client: Anthropic, themes_result: dict, ordered_quarters):
    requests = []
    manifest = {}
    for later_item, earlier_item in zip(ordered_quarters[1:], ordered_quarters[:-1]):
        delta_key = f"{later_item['label']}_vs_{earlier_item['label']}"
        ranked_themes = themes_result.get(later_item["period"], {}).get("themes", [])
        if not ranked_themes:
            continue
        merged_id = f"{delta_key}__themes"
        merged_themes = COMMON_THEME_SET + ranked_themes
        requests.append(build_theme_delta_batch_request(merged_id, merged_themes, later_item, earlier_item))
        manifest[delta_key] = {
            "merged_id": merged_id,
            "ranked_themes": ranked_themes,
        }
    batch = client.messages.batches.create(requests=requests)
    return batch.id, manifest


def split_merged_result(merged_result: dict):
    entries = merged_result.get("historical_analysis", []) if merged_result else []
    readthrough_result = {
        "historical_analysis": [
            {
                "period": e["period"],
                "signal": e["signal"],
                "read_through": e["read_through"],
                "rationale": e["rationale"],
            }
            for e in entries
        ]
    }
    bullbear_result = {
        "historical_analysis": [
            {"period": e["period"], "bull": e["bull"], "bear": e["bear"]}
            for e in entries
        ]
    }
    return readthrough_result, bullbear_result


def split_theme_delta_result(merged_result: dict, ranked_themes: list):
    ranked_set = set(ranked_themes)
    entries = merged_result.get("theme_analysis", []) if merged_result else []
    common_result = {"theme_analysis": [e for e in entries if e["theme"] not in ranked_set]}
    ranked_result = {"theme_analysis": [e for e in entries if e["theme"] in ranked_set]}
    return common_result, ranked_result


def fetch_raw_batch_results(client: Anthropic, batch_id: str) -> dict:
    results = {}
    for entry in client.messages.batches.results(batch_id):
        if entry.result.type == "succeeded":
            content = entry.result.message.content
            text = "".join(block.text for block in content if block.type == "text")
            block_summary = [{"type": block.type} for block in content]
            results[entry.custom_id] = (text, None, block_summary)
        else:
            error_message = getattr(entry.result, "error", None)
            error_str = str(error_message) if error_message is not None else entry.result.type
            results[entry.custom_id] = (None, error_str, None)
    return results


# ============================================================
# DB EXPORT
# ============================================================

def find_ticker_worksheet(workbook: openpyxl.Workbook, ticker: str):
    for name in workbook.sheetnames:
        if name.strip().upper() == ticker.strip().upper():
            return name
    return None


DB_PERIOD_COLUMN = 1
DB_FIRST_DATA_ROW = 6


def period_to_db_format(period: str) -> str:
    quarter, year = period.split()
    return f"{year[-2:]}Q{quarter[1]}"


def find_period_row(worksheet, period: str) -> int | None:
    target = period_to_db_format(period)
    for row in worksheet.iter_rows(min_row=DB_FIRST_DATA_ROW, min_col=DB_PERIOD_COLUMN, max_col=DB_PERIOD_COLUMN):
        cell = row[0]
        value = cell.value
        if value is not None and str(value).strip().upper() == target:
            return cell.row
    return None


DB_READTHROUGH_COL = column_index_from_string("DX")
DB_BULL_COL = DB_READTHROUGH_COL + 1
DB_BEAR_COL = DB_READTHROUGH_COL + 2
DB_COMMON_THEME_COLUMNS = {
    COMMON_THEME_SET[0]: (DB_READTHROUGH_COL + 3, DB_READTHROUGH_COL + 4),
    COMMON_THEME_SET[1]: (DB_READTHROUGH_COL + 5, DB_READTHROUGH_COL + 6),
    COMMON_THEME_SET[2]: (DB_READTHROUGH_COL + 7, DB_READTHROUGH_COL + 8),
}
DB_VARIABLE_THEME_START_COL = DB_READTHROUGH_COL + 9
DB_VARIABLE_THEME_HEADER_ROW = 4


def format_signal_cell(signal: str, text: str, rationale: str) -> str:
    return f"{signal} {text} — {rationale}"


def build_delta_lookup(result: dict) -> dict:
    return {entry["period"]: entry for entry in result.get("historical_analysis", [])}


def get_or_create_variable_theme_columns(worksheet, theme_name: str):
    col = DB_VARIABLE_THEME_START_COL
    while True:
        header_cell = worksheet.cell(row=DB_VARIABLE_THEME_HEADER_ROW, column=col)
        if header_cell.value is None:
            header_cell.value = theme_name
            return col, col + 1, col + 2
        if str(header_cell.value).strip() == theme_name.strip():
            return col, col + 1, col + 2
        col += 3


def get_existing_variable_themes(worksheet) -> list[str]:
    themes = []
    col = DB_VARIABLE_THEME_START_COL
    while True:
        cell = worksheet.cell(row=DB_VARIABLE_THEME_HEADER_ROW, column=col)
        if cell.value is None:
            break
        themes.append(str(cell.value).strip())
        col += 3
    return themes


def write_db_row(
    worksheet,
    row: int,
    readthrough_entry: dict,
    bullbear_entry: dict,
    theme_common_result: dict,
    theme_variable_result: dict = None,
    ranked_themes: list = None,
):
    if readthrough_entry is not None:
        worksheet.cell(row=row, column=DB_READTHROUGH_COL).value = format_signal_cell(
            readthrough_entry["signal"], readthrough_entry["read_through"], readthrough_entry["rationale"]
        )

    if bullbear_entry is not None:
        worksheet.cell(row=row, column=DB_BULL_COL).value = format_signal_cell(
            bullbear_entry["bull"]["topic"], bullbear_entry["bull"]["expectation"], bullbear_entry["bull"]["context"]
        )
        worksheet.cell(row=row, column=DB_BEAR_COL).value = format_signal_cell(
            bullbear_entry["bear"]["topic"], bullbear_entry["bear"]["expectation"], bullbear_entry["bear"]["context"]
        )

    if theme_common_result is not None:
        theme_entries = {entry["theme"]: entry for entry in theme_common_result.get("theme_analysis", [])}
        for theme_name, (mgmt_col, analyst_col) in DB_COMMON_THEME_COLUMNS.items():
            entry = theme_entries.get(theme_name)
            if entry is None:
                continue
            worksheet.cell(row=row, column=mgmt_col).value = format_signal_cell(
                entry["mgmt"]["signal"], entry["mgmt"]["message"], entry["mgmt"]["rationale"]
            )
            worksheet.cell(row=row, column=analyst_col).value = format_signal_cell(
                entry["analyst"]["signal"], entry["analyst"]["tone"], entry["analyst"]["rationale"]
            )

    if theme_variable_result is not None:
        for entry in theme_variable_result.get("theme_analysis", []):
            theme_name = entry["theme"]
            rank_col, mgmt_col, analyst_col = get_or_create_variable_theme_columns(worksheet, theme_name)
            if ranked_themes and theme_name in ranked_themes:
                worksheet.cell(row=row, column=rank_col).value = ranked_themes.index(theme_name) + 1
            worksheet.cell(row=row, column=mgmt_col).value = format_signal_cell(
                entry["mgmt"]["signal"], entry["mgmt"]["message"], entry["mgmt"]["rationale"]
            )
            worksheet.cell(row=row, column=analyst_col).value = format_signal_cell(
                entry["analyst"]["signal"], entry["analyst"]["tone"], entry["analyst"]["rationale"]
            )


def write_all_theme_results(
    worksheet,
    parsed_quarters,
    period_rows: dict,
    themes: dict,
    readthrough_lookup: dict,
    bullbear_lookup: dict,
    theme_delta_by_key: dict,
):
    for later_item, _ in reversed(list(zip(parsed_quarters[1:], parsed_quarters[:-1]))):
        for theme_name in themes.get(later_item["period"], {}).get("themes", []):
            get_or_create_variable_theme_columns(worksheet, theme_name)

    for later_item, earlier_item in zip(parsed_quarters[1:], parsed_quarters[:-1]):
        delta_key = f"{later_item['label']}_vs_{earlier_item['label']}"
        entry = theme_delta_by_key.get(delta_key)
        if entry is None:
            continue
        write_db_row(
            worksheet,
            period_rows[later_item["period"]],
            readthrough_lookup.get(delta_key),
            bullbear_lookup.get(delta_key),
            entry["common"],
            entry["ranked"],
            entry["ranked_themes"],
        )


# ============================================================
# UTILITIES
# ============================================================

def parse_md_metadata(filename: str):
    name = filename.rsplit(".", 1)[0]
    parts = name.split("-")
    ticker = " ".join(parts[:2]) if len(parts) >= 2 else None
    period_match = re.search(r"Q[1-4]\s+20\d{2}", name)
    period = period_match.group(0) if period_match else None
    return ticker, period


def quarter_sort_key(period: str):
    quarter, year = period.split()
    return int(year), int(quarter[1])


# ============================================================
# STREAMLIT UI
# ============================================================

st.set_page_config(page_title="Cross-Quarter Theme Analysis", layout="wide")
st.title("Cross-Quarter Theme Analysis")

with st.sidebar:
    api_key = st.text_input("Anthropic API key", type="password")

theme_md_files = st.file_uploader(
    "Upload telegraphic .md files (one company, 2 or more consecutive quarters)",
    type=["md"],
    accept_multiple_files=True,
)

db_file = st.file_uploader(
    "Upload DB format.xlsx (workbook to update with this analysis)",
    type=["xlsx"],
)

if theme_md_files:
    parsed_quarters = []
    for f in theme_md_files:
        ticker, period = parse_md_metadata(f.name)
        if period is None:
            st.error(f"{f.name}: could not parse a quarter/period (e.g. 'Q1 2026') from filename")
            continue
        parsed_quarters.append(
            {
                "filename": f.name,
                "ticker": ticker,
                "period": period,
                "content": f.read().decode("utf-8"),
            }
        )

    if len(parsed_quarters) < MIN_QUARTERS:
        st.warning(f"Need at least {MIN_QUARTERS} valid .md files; found {len(parsed_quarters)}.")
    else:
        parsed_quarters.sort(key=lambda item: quarter_sort_key(item["period"]))
        labels = build_quarter_labels(len(parsed_quarters))
        for label, item in zip(labels, parsed_quarters):
            item["label"] = label

        st.write({item["label"]: f"{item['ticker']} {item['period']}" for item in parsed_quarters})

        ticker = parsed_quarters[0]["ticker"]
        sheet_name = None
        period_rows = {}
        if db_file:
            workbook = openpyxl.load_workbook(db_file)
            sheet_name = find_ticker_worksheet(workbook, ticker)
            if sheet_name is None:
                st.error(f"No tab found matching ticker '{ticker}' in {db_file.name}")
            else:
                st.success(f"Found matching tab: '{sheet_name}'")
                worksheet = workbook[sheet_name]
                missing_periods = []
                for item in parsed_quarters:
                    row = find_period_row(worksheet, item["period"])
                    if row is None:
                        missing_periods.append(item["period"])
                    else:
                        period_rows[item["period"]] = row
                if missing_periods:
                    st.error(
                        f"No row found in '{sheet_name}' for: {', '.join(missing_periods)} "
                        f"(expected column A format like '{period_to_db_format(parsed_quarters[0]['period'])}')"
                    )
                else:
                    st.success(
                        "Matched rows: "
                        + ", ".join(f"{period} -> row {row}" for period, row in period_rows.items())
                    )
        else:
            st.info(f"Upload DB format.xlsx above to update the '{ticker}' tab with this analysis.")

        rows_ready = db_file and sheet_name and len(period_rows) == len(parsed_quarters)

        existing_themes = get_existing_variable_themes(worksheet) if rows_ready else []

        use_batch = st.checkbox(
            "Use batch processing (cheaper, ~50% cost)",
            help="Submits the analysis as Batches API jobs at ~50% lower API cost. "
            "Runs automatically: stage 1 (themes, read-through, bull/bear) runs first, "
            "then stage 2 (per-delta theme analysis) is submitted and awaited. "
            "Keep this tab open until the download button appears.",
        )

        if use_batch:
            if st.button("Run batch", disabled=not (api_key and rows_ready)):
                client = Anthropic(api_key=api_key)
                status = st.empty()

                status.info("Submitting stage 1 batch (themes, read-through, bull/bear)...")
                batch_id = submit_stage1_batch(client, parsed_quarters, existing_themes)

                while True:
                    batch = client.messages.batches.retrieve(batch_id)
                    if batch.processing_status == "ended":
                        break
                    counts = batch.request_counts
                    status.info(
                        f"Stage 1 running — processing: {counts.processing}, "
                        f"succeeded: {counts.succeeded}, errored: {counts.errored}. Checking again in 30s..."
                    )
                    time.sleep(30)

                status.info("Fetching stage 1 results...")
                raw_results = fetch_raw_batch_results(client, batch_id)

                themes_text, themes_err, _ = raw_results.get("themes", (None, "missing", None))
                themes_result = None
                if themes_text:
                    themes_result, parse_err = parse_themes_response(themes_text)
                    if parse_err:
                        st.error(f"themes: {parse_err}")
                        with st.expander("themes raw response"):
                            st.text(themes_text)
                else:
                    st.error(f"themes batch error: {themes_err}")

                readthrough_result = None
                bullbear_result = None
                merged_text, merged_err, merged_blocks = raw_results.get("merged", (None, "missing", None))
                if merged_text is not None:
                    if not merged_text:
                        st.error(f"merged: batch succeeded but response contained no text (blocks: {merged_blocks})")
                    else:
                        merged_result, merged_parse_err = parse_fenced_json_response(merged_text)
                        if merged_parse_err:
                            st.error(f"merged: JSON parse failed — {merged_parse_err}")
                            with st.expander("merged raw response"):
                                st.text(merged_text)
                        else:
                            readthrough_result, bullbear_result = split_merged_result(merged_result)
                else:
                    st.error(f"merged batch error: {merged_err}")

                if themes_result is not None:
                    st.json(themes_result)

                    status.info("Submitting stage 2 batch (per-delta theme analysis)...")
                    stage2_id, manifest = submit_stage2_batch(client, themes_result, parsed_quarters)

                    while True:
                        batch = client.messages.batches.retrieve(stage2_id)
                        if batch.processing_status == "ended":
                            break
                        counts = batch.request_counts
                        status.info(
                            f"Stage 2 running — processing: {counts.processing}, "
                            f"succeeded: {counts.succeeded}, errored: {counts.errored}. Checking again in 30s..."
                        )
                        time.sleep(30)

                    status.info("Fetching stage 2 results and writing workbook...")
                    raw_results = fetch_raw_batch_results(client, stage2_id)
                    readthrough_lookup = build_delta_lookup(readthrough_result) if readthrough_result else {}
                    bullbear_lookup = build_delta_lookup(bullbear_result) if bullbear_result else {}

                    theme_delta_by_key = {}
                    for delta_key, ids in manifest.items():
                        merged_text, merged_err, merged_blocks = raw_results.get(ids["merged_id"], (None, "missing", None))
                        merged_result = None
                        if merged_text is not None:
                            if not merged_text:
                                st.error(f"{delta_key}: batch succeeded but response contained no text (blocks: {merged_blocks})")
                            else:
                                merged_result, merged_parse_err = parse_fenced_json_response(merged_text)
                                if merged_parse_err:
                                    st.error(f"{delta_key}: JSON parse failed — {merged_parse_err}")
                                    with st.expander(f"{delta_key} raw response"):
                                        st.text(merged_text)
                        else:
                            st.error(f"{delta_key} batch error: {merged_err}")

                        common_result, ranked_result = split_theme_delta_result(
                            merged_result, ids["ranked_themes"]
                        )
                        theme_delta_by_key[delta_key] = {
                            "common": common_result,
                            "ranked": ranked_result,
                            "ranked_themes": ids["ranked_themes"],
                        }

                    write_all_theme_results(
                        worksheet, parsed_quarters, period_rows, themes_result,
                        readthrough_lookup, bullbear_lookup, theme_delta_by_key,
                    )

                    st.json(readthrough_result)
                    st.json(bullbear_result)
                    st.json(theme_delta_by_key)

                    output_buffer = io.BytesIO()
                    workbook.save(output_buffer)
                    status.success("Batch analysis complete.")
                    st.download_button(
                        label=f"Download updated {db_file.name}",
                        data=output_buffer.getvalue(),
                        file_name=db_file.name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="batch_dl",
                    )
                else:
                    status.error("Stage 1 failed to produce themes — cannot continue.")
            elif not api_key:
                st.info("Enter your Anthropic API key in the sidebar to begin.")

        elif st.button("Analyze Themes", disabled=not (api_key and rows_ready)):
            client = Anthropic(api_key=api_key)
            with st.spinner("Analyzing cross-quarter themes..."):
                themes, error = analyze_themes(client, parsed_quarters, existing_themes)
            if error:
                st.error(error)
            else:
                st.session_state["theme_analysis_result"] = themes
                st.json(themes)

                with st.spinner("Analyzing sector read-through and bull/bear signals..."):
                    merged, merged_error = analyze_merged(client, parsed_quarters)
                readthrough_lookup = {}
                bullbear_lookup = {}
                if merged_error:
                    st.error(merged_error)
                else:
                    readthrough, bullbear = split_merged_result(merged)
                    st.session_state["readthrough_analysis_result"] = readthrough
                    st.json(readthrough)
                    readthrough_lookup = build_delta_lookup(readthrough)
                    st.session_state["bullbear_analysis_result"] = bullbear
                    st.json(bullbear)
                    bullbear_lookup = build_delta_lookup(bullbear)

                theme_delta_by_key = {}
                for later_item, earlier_item in zip(parsed_quarters[1:], parsed_quarters[:-1]):
                    delta_key = f"{later_item['label']}_vs_{earlier_item['label']}"
                    ranked_themes = themes.get(later_item["period"], {}).get("themes", [])
                    if not ranked_themes:
                        st.error(f"{delta_key}: no ranked themes found for {later_item['period']}")
                        continue

                    merged_themes = COMMON_THEME_SET + ranked_themes
                    with st.spinner(f"Analyzing themes for {delta_key}..."):
                        merged_result, merged_error = analyze_theme_delta(
                            client, merged_themes, later_item, earlier_item
                        )

                    if merged_error:
                        st.error(f"{delta_key}: {merged_error}")

                    common_result, ranked_result = split_theme_delta_result(merged_result, ranked_themes)

                    theme_delta_by_key[delta_key] = {
                        "common": common_result,
                        "ranked": ranked_result,
                        "ranked_themes": ranked_themes,
                    }
                    st.write(f"Theme delta analysis — {delta_key}")
                    st.json(theme_delta_by_key[delta_key])

                st.session_state["theme_delta_analysis_result"] = theme_delta_by_key

                write_all_theme_results(
                    worksheet, parsed_quarters, period_rows, themes,
                    readthrough_lookup, bullbear_lookup, theme_delta_by_key,
                )

                output_buffer = io.BytesIO()
                workbook.save(output_buffer)
                st.download_button(
                    label=f"Download updated {db_file.name}",
                    data=output_buffer.getvalue(),
                    file_name=db_file.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
elif not api_key:
    st.info("Enter your Anthropic API key in the sidebar to begin.")
else:
    st.info(f"Upload at least {MIN_QUARTERS} telegraphic .md files to begin.")
