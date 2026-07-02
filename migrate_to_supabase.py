"""
One-time migration script: reads the existing Database.xlsx and populates
all Supabase tables (tickers, quarters, quarter_signals, variable_themes,
external_data).

Usage:
    python migrate_to_supabase.py
    DATABASE_URL must be set as an environment variable, or pass it as first arg:
    python migrate_to_supabase.py "postgresql://postgres:PASSWORD@db.REF.supabase.co:5432/postgres"
"""

import sys
import openpyxl
import db

EXCEL_PATH = r"C:\Users\charles.yang\Earnings-Score2\Database.xlsx"


def migrate_signals(conn, workbook, quarter_id_map: dict):
    """Read DX+ columns from every sheet and write to quarter_signals + variable_themes."""
    cur = conn.cursor()

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ticker = str(ws.cell(1, 2).value or "").strip()
        if ticker not in quarter_id_map:
            continue

        var_themes = db.get_variable_themes_from_sheet(ws)
        print(f"  {ticker}: {len(var_themes)} variable theme column(s)")

        for row in range(db.DB_FIRST_DATA_ROW, ws.max_row + 1):
            quarter_val = ws.cell(row, db.DB_PERIOD_COLUMN).value
            if quarter_val is None:
                break
            quarter = str(quarter_val).strip()
            quarter_id = quarter_id_map[ticker].get(quarter)
            if quarter_id is None:
                continue

            # Skip rows with no analysis data
            if ws.cell(row, db.DB_READTHROUGH_COL).value is None:
                continue

            signals = db._blank_signals()
            signals["quarter_id"] = quarter_id

            rt_sig, rt_text, rt_rat = db.parse_excel_signal_cell(ws.cell(row, db.DB_READTHROUGH_COL).value)
            signals["read_through_signal"]      = rt_sig
            signals["read_through_description"] = db._combine_description(rt_text, rt_rat)

            bull_sig, bull_text, bull_ctx = db.parse_excel_signal_cell(ws.cell(row, db.DB_BULL_COL).value)
            signals["bull_signal"]      = bull_sig
            signals["bull_description"] = db._combine_description(bull_text, bull_ctx)

            bear_sig, bear_text, bear_ctx = db.parse_excel_signal_cell(ws.cell(row, db.DB_BEAR_COL).value)
            signals["bear_signal"]      = bear_sig
            signals["bear_description"] = db._combine_description(bear_text, bear_ctx)

            for theme_name, (mgmt_col, analyst_col) in db.DB_COMMON_THEME_COLUMNS.items():
                prefix = db.COMMON_THEME_PREFIX[theme_name]
                m_sig, m_msg, m_rat = db.parse_excel_signal_cell(ws.cell(row, mgmt_col).value)
                a_sig, a_tone, a_rat = db.parse_excel_signal_cell(ws.cell(row, analyst_col).value)
                signals[f"{prefix}_mgmt_signal"]         = m_sig
                signals[f"{prefix}_mgmt_description"]    = db._combine_description(m_msg, m_rat)
                signals[f"{prefix}_analyst_signal"]      = a_sig
                signals[f"{prefix}_analyst_description"] = db._combine_description(a_tone, a_rat)

            db.upsert_quarter_signals(cur, quarter_id, signals)

            for theme_name, start_col in var_themes:
                rank_val  = ws.cell(row, start_col).value
                mgmt_val  = ws.cell(row, start_col + 1).value
                anal_val  = ws.cell(row, start_col + 2).value
                if rank_val is None and mgmt_val is None and anal_val is None:
                    continue
                rank = int(rank_val) if rank_val is not None else None
                m_sig, m_msg, m_rat = db.parse_excel_signal_cell(mgmt_val)
                a_sig, a_tone, a_rat = db.parse_excel_signal_cell(anal_val)
                db.upsert_variable_theme(
                    cur, quarter_id, theme_name, rank,
                    m_sig, db._combine_description(m_msg, m_rat),
                    a_sig, db._combine_description(a_tone, a_rat),
                )

        conn.commit()
        print(f"  {ticker}: done")

    cur.close()


def main():
    database_url = sys.argv[1] if len(sys.argv) > 1 else None

    print(f"Opening {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    print("Connecting to Supabase...")
    conn = db.get_connection(database_url)

    print("\nPhase 1: syncing tickers, quarters, and Bloomberg data...")
    quarter_id_map = db.sync_workbook_structure(conn, wb)
    print(f"  Synced {len(quarter_id_map)} tickers.")

    print("\nPhase 2: migrating analysis signals and themes...")
    migrate_signals(conn, wb, quarter_id_map)

    conn.close()
    print("\nMigration complete.")


if __name__ == "__main__":
    main()
