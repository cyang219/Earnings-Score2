import json
import os
import re
import time
from typing import Optional

import openpyxl
import psycopg2
import streamlit as st
from anthropic import Anthropic
from openpyxl.utils import column_index_from_string, get_column_letter

MODEL = "claude-sonnet-4-6"
THEMES_EFFORT = "medium"
READTHROUGH_MODEL = "claude-sonnet-4-6"
READTHROUGH_EFFORT = "medium"
THEME_DELTA_MODEL = "claude-sonnet-4-6"
THEME_DELTA_EFFORT = "low"

MIN_QUARTERS = 2


# ============================================================
# DATABASE CONSTANTS
# ============================================================

COMMON_THEME_SET = [
    "Top-line Growth (e.g. Revenue, Bookings, GMV, ARR, etc.)",
    "Bottom-line Expansion/Contraction (e.g. Margins, Return on Investment, Take-rate, etc.)",
    "Financial Guidance and Forward-looking Business Commentary",
]

COMMON_THEME_PREFIX = {
    COMMON_THEME_SET[0]: "topline",
    COMMON_THEME_SET[1]: "bottomline",
    COMMON_THEME_SET[2]: "guidance",
}

DB_PERIOD_COLUMN = 1
DB_DATE_COLUMN = 2
DB_FIRST_DATA_ROW = 6
DB_READTHROUGH_COL = column_index_from_string("DX")       # 128
DB_BULL_COL = DB_READTHROUGH_COL + 1
DB_BEAR_COL = DB_READTHROUGH_COL + 2
DB_COMMON_THEME_COLUMNS = {
    COMMON_THEME_SET[0]: (DB_READTHROUGH_COL + 3, DB_READTHROUGH_COL + 4),
    COMMON_THEME_SET[1]: (DB_READTHROUGH_COL + 5, DB_READTHROUGH_COL + 6),
    COMMON_THEME_SET[2]: (DB_READTHROUGH_COL + 7, DB_READTHROUGH_COL + 8),
}
DB_VARIABLE_THEME_START_COL = DB_READTHROUGH_COL + 9      # 137
DB_VARIABLE_THEME_HEADER_ROW = 4
DB_BLOOMBERG_END_COL = column_index_from_string("DW")     # 127


# ============================================================
# DATABASE CONNECTION
# ============================================================

def get_connection(database_url: str = None) -> psycopg2.extensions.connection:
    return psycopg2.connect(database_url or os.environ["DATABASE_URL"])


# ============================================================
# SIGNAL PARSING
# ============================================================

def signal_to_int(signal_str: str) -> Optional[int]:
    """'↑:' / '↓:' / '→:' → 1 / -1 / 0."""
    if not signal_str:
        return None
    s = str(signal_str).strip()
    if s.startswith("↑"):
        return 1
    if s.startswith("↓"):
        return -1
    if s.startswith("→"):
        return 0
    return None


def parse_excel_signal_cell(value) -> tuple:
    """Parse '↑: text — rationale' from an Excel cell into (signal_int, text, rationale)."""
    if value is None:
        return None, None, None
    text = str(value).strip()
    match = re.match(r"^([↑↓→]):\s*(.*)", text, flags=re.DOTALL)
    if not match:
        return None, text, None
    signal = signal_to_int(match.group(1))
    rest = match.group(2).strip()
    parts = rest.split(" — ", 1)
    description = parts[0].strip() if parts else None
    rationale = parts[1].strip() if len(parts) > 1 else None
    return signal, description, rationale


# ============================================================
# PERIOD FORMAT
# ============================================================

def period_to_quarter(period: str) -> str:
    """'Q1 2026' → '26Q1'"""
    quarter, year = period.split()
    return f"{year[-2:]}Q{quarter[1]}"


# ============================================================
# EXCEL HELPERS
# ============================================================

def build_bloomberg_header_map(worksheet) -> dict:
    """Column index → display name for Bloomberg columns C (3) through DW (127).
    Uses row 4 as primary header, falls back to row 1 (Bloomberg field code)."""
    headers = {}
    for col in range(3, DB_BLOOMBERG_END_COL + 1):
        h4 = worksheet.cell(4, col).value
        h1 = worksheet.cell(1, col).value
        if h4 and str(h4).strip():
            headers[col] = str(h4).strip()
        elif h1 and str(h1).strip():
            headers[col] = str(h1).strip()
    return headers


def get_variable_themes_from_sheet(worksheet) -> list:
    """Return [(theme_name, start_col)] for variable theme columns in this sheet."""
    themes = []
    col = DB_VARIABLE_THEME_START_COL
    while True:
        v = worksheet.cell(DB_VARIABLE_THEME_HEADER_ROW, col).value
        if v is None:
            break
        themes.append((str(v).strip(), col))
        col += 3
    return themes


# ============================================================
# UPSERT FUNCTIONS
# ============================================================

def upsert_ticker(cur, ticker: str, index_name: str, currency: str, sector: str):
    cur.execute("""
        INSERT INTO tickers (ticker, index_name, currency, sector)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (ticker) DO UPDATE SET
            index_name = EXCLUDED.index_name,
            currency   = EXCLUDED.currency,
            sector     = EXCLUDED.sector
    """, (ticker, index_name, currency, sector))


def upsert_quarter(cur, ticker: str, quarter: str, date) -> int:
    date_val = date.date() if hasattr(date, "date") else date
    cur.execute("""
        INSERT INTO quarters (ticker, quarter, date)
        VALUES (%s, %s, %s)
        ON CONFLICT (ticker, quarter) DO UPDATE SET date = EXCLUDED.date
        RETURNING id
    """, (ticker, quarter, date_val))
    return cur.fetchone()[0]


def get_quarter_id(cur, ticker: str, quarter: str) -> Optional[int]:
    cur.execute(
        "SELECT id FROM quarters WHERE ticker = %s AND quarter = %s",
        (ticker, quarter),
    )
    row = cur.fetchone()
    return row[0] if row else None


def upsert_quarter_signals(cur, quarter_id: int, data: dict):
    cur.execute("""
        INSERT INTO quarter_signals (
            quarter_id,
            read_through_signal, read_through_description,
            bull_signal, bull_description,
            bear_signal, bear_description,
            topline_mgmt_signal, topline_mgmt_description,
            topline_analyst_signal, topline_analyst_description,
            bottomline_mgmt_signal, bottomline_mgmt_description,
            bottomline_analyst_signal, bottomline_analyst_description,
            guidance_mgmt_signal, guidance_mgmt_description,
            guidance_analyst_signal, guidance_analyst_description
        ) VALUES (
            %(quarter_id)s,
            %(read_through_signal)s, %(read_through_description)s,
            %(bull_signal)s, %(bull_description)s,
            %(bear_signal)s, %(bear_description)s,
            %(topline_mgmt_signal)s, %(topline_mgmt_description)s,
            %(topline_analyst_signal)s, %(topline_analyst_description)s,
            %(bottomline_mgmt_signal)s, %(bottomline_mgmt_description)s,
            %(bottomline_analyst_signal)s, %(bottomline_analyst_description)s,
            %(guidance_mgmt_signal)s, %(guidance_mgmt_description)s,
            %(guidance_analyst_signal)s, %(guidance_analyst_description)s
        )
        ON CONFLICT (quarter_id) DO UPDATE SET
            read_through_signal            = EXCLUDED.read_through_signal,
            read_through_description       = EXCLUDED.read_through_description,
            bull_signal                    = EXCLUDED.bull_signal,
            bull_description               = EXCLUDED.bull_description,
            bear_signal                    = EXCLUDED.bear_signal,
            bear_description               = EXCLUDED.bear_description,
            topline_mgmt_signal            = EXCLUDED.topline_mgmt_signal,
            topline_mgmt_description       = EXCLUDED.topline_mgmt_description,
            topline_analyst_signal         = EXCLUDED.topline_analyst_signal,
            topline_analyst_description    = EXCLUDED.topline_analyst_description,
            bottomline_mgmt_signal         = EXCLUDED.bottomline_mgmt_signal,
            bottomline_mgmt_description    = EXCLUDED.bottomline_mgmt_description,
            bottomline_analyst_signal      = EXCLUDED.bottomline_analyst_signal,
            bottomline_analyst_description = EXCLUDED.bottomline_analyst_description,
            guidance_mgmt_signal           = EXCLUDED.guidance_mgmt_signal,
            guidance_mgmt_description      = EXCLUDED.guidance_mgmt_description,
            guidance_analyst_signal        = EXCLUDED.guidance_analyst_signal,
            guidance_analyst_description   = EXCLUDED.guidance_analyst_description
    """, {"quarter_id": quarter_id, **data})


def _combine_description(text: Optional[str], rationale: Optional[str]) -> Optional[str]:
    if text and rationale:
        return f"{text} — {rationale}"
    return text or rationale


def upsert_variable_theme(
    cur,
    quarter_id: int,
    theme_name: str,
    rank: Optional[int],
    mgmt_signal: Optional[int],
    mgmt_description: Optional[str],
    analyst_signal: Optional[int],
    analyst_description: Optional[str],
):
    cur.execute("""
        INSERT INTO variable_themes (
            quarter_id, theme_name, rank,
            mgmt_signal, mgmt_description,
            analyst_signal, analyst_description
        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (quarter_id, theme_name) DO UPDATE SET
            rank                = EXCLUDED.rank,
            mgmt_signal         = EXCLUDED.mgmt_signal,
            mgmt_description    = EXCLUDED.mgmt_description,
            analyst_signal      = EXCLUDED.analyst_signal,
            analyst_description = EXCLUDED.analyst_description
    """, (
        quarter_id, theme_name, rank,
        mgmt_signal, mgmt_description,
        analyst_signal, analyst_description,
    ))


def upsert_external_data(cur, ticker: str, quarter: str, data: dict):
    cur.execute("""
        INSERT INTO external_data (ticker, quarter, data)
        VALUES (%s, %s, %s::jsonb)
        ON CONFLICT (ticker, quarter) DO UPDATE SET
            data = external_data.data || EXCLUDED.data
    """, (ticker, quarter, json.dumps(data)))


def get_existing_variable_themes(cur, ticker: str) -> list:
    cur.execute("""
        SELECT DISTINCT vt.theme_name
        FROM variable_themes vt
        JOIN quarters q ON q.id = vt.quarter_id
        WHERE q.ticker = %s
        ORDER BY vt.theme_name
    """, (ticker,))
    return [row[0] for row in cur.fetchall()]


# ============================================================
# EXCEL → SUPABASE: STRUCTURE SYNC
# ============================================================

def sync_workbook_structure(conn, workbook) -> dict:
    """Upsert tickers, quarters, and Bloomberg external_data from the uploaded workbook.

    Returns {ticker: {quarter_str: quarter_id}} for all sheets processed.
    """
    cur = conn.cursor()
    quarter_id_map = {}

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ticker = str(ws.cell(1, 2).value or "").strip()
        if not ticker:
            continue

        index_name = str(ws.cell(1, 1).value or "").strip() or None
        currency   = str(ws.cell(2, 1).value or "").strip() or None
        sector     = str(ws.cell(2, 2).value or "").strip() or None

        upsert_ticker(cur, ticker, index_name, currency, sector)

        bloomberg_headers = build_bloomberg_header_map(ws)
        quarter_id_map[ticker] = {}

        for row in range(DB_FIRST_DATA_ROW, ws.max_row + 1):
            quarter_val = ws.cell(row, DB_PERIOD_COLUMN).value
            if quarter_val is None:
                break
            quarter = str(quarter_val).strip()
            date_val = ws.cell(row, DB_DATE_COLUMN).value
            qid = upsert_quarter(cur, ticker, quarter, date_val)
            quarter_id_map[ticker][quarter] = qid

            bloomberg_data = {}
            for col, header in bloomberg_headers.items():
                val = ws.cell(row, col).value
                if val is None:
                    continue
                str_val = str(val).strip()
                if str_val in ("", ", "):
                    continue
                bloomberg_data[header] = val.isoformat() if hasattr(val, "isoformat") else val
            if bloomberg_data:
                upsert_external_data(cur, ticker, quarter, bloomberg_data)

    conn.commit()
    cur.close()
    return quarter_id_map


# ============================================================
# ANALYSIS RESULTS → SUPABASE
# ============================================================

def _blank_signals() -> dict:
    """Return a dict with all quarter_signals fields set to None."""
    fields = [
        "read_through_signal", "read_through_description",
        "bull_signal", "bull_description",
        "bear_signal", "bear_description",
    ]
    for prefix in ("topline", "bottomline", "guidance"):
        fields += [
            f"{prefix}_mgmt_signal", f"{prefix}_mgmt_description",
            f"{prefix}_analyst_signal", f"{prefix}_analyst_description",
        ]
    return {f: None for f in fields}


def write_analysis_results(
    conn,
    ticker: str,
    quarter_id_map: dict,
    parsed_quarters: list,
    readthrough_result: Optional[dict],
    bullbear_result: Optional[dict],
    theme_delta_by_key: dict,
):
    """Write theme analysis results to quarter_signals and variable_themes."""
    cur = conn.cursor()

    rt_lookup = {e["period"]: e for e in (readthrough_result or {}).get("historical_analysis", [])}
    bb_lookup = {e["period"]: e for e in (bullbear_result or {}).get("historical_analysis", [])}

    for later_item, earlier_item in zip(parsed_quarters[1:], parsed_quarters[:-1]):
        delta_key = f"{later_item['label']}_vs_{earlier_item['label']}"
        later_quarter = period_to_quarter(later_item["period"])
        quarter_id = quarter_id_map.get(later_quarter)
        if quarter_id is None:
            continue

        rt  = rt_lookup.get(delta_key)
        bb  = bb_lookup.get(delta_key)
        entry = theme_delta_by_key.get(delta_key)

        signals = _blank_signals()
        signals["quarter_id"] = quarter_id

        if rt:
            signals["read_through_signal"]      = signal_to_int(rt.get("signal"))
            signals["read_through_description"] = _combine_description(rt.get("read_through"), rt.get("rationale"))

        if bb:
            signals["bull_signal"]      = signal_to_int(bb["bull"].get("topic"))
            signals["bull_description"] = _combine_description(bb["bull"].get("expectation"), bb["bull"].get("context"))
            signals["bear_signal"]      = signal_to_int(bb["bear"].get("topic"))
            signals["bear_description"] = _combine_description(bb["bear"].get("expectation"), bb["bear"].get("context"))

        if entry:
            common_entries = {e["theme"]: e for e in entry["common"].get("theme_analysis", [])}
            for theme_name, prefix in COMMON_THEME_PREFIX.items():
                e = common_entries.get(theme_name)
                if not e:
                    continue
                signals[f"{prefix}_mgmt_signal"]          = signal_to_int(e["mgmt"].get("signal"))
                signals[f"{prefix}_mgmt_description"]     = _combine_description(e["mgmt"].get("message"), e["mgmt"].get("rationale"))
                signals[f"{prefix}_analyst_signal"]       = signal_to_int(e["analyst"].get("signal"))
                signals[f"{prefix}_analyst_description"]  = _combine_description(e["analyst"].get("tone"), e["analyst"].get("rationale"))

        upsert_quarter_signals(cur, quarter_id, signals)

        if entry:
            ranked_themes = entry.get("ranked_themes", [])
            for vt in entry["ranked"].get("theme_analysis", []):
                theme_name = vt["theme"]
                rank = (ranked_themes.index(theme_name) + 1) if theme_name in ranked_themes else None
                upsert_variable_theme(
                    cur, quarter_id, theme_name, rank,
                    signal_to_int(vt["mgmt"].get("signal")),
                    _combine_description(vt["mgmt"].get("message"), vt["mgmt"].get("rationale")),
                    signal_to_int(vt["analyst"].get("signal")),
                    _combine_description(vt["analyst"].get("tone"), vt["analyst"].get("rationale")),
                )

    conn.commit()
    cur.close()


# ============================================================
# PROMPTS
# ============================================================

def build_quarter_labels(n: int) -> list[str]:
    return [f"Q-{n - 1 - i}" for i in range(n - 1)] + ["Latest_Q"]


def _oxford_join(items: list[str]) -> str:
    if len(items) == 1:
        return items[0]
    if len(items) == 2:
        return f"{items[0]} and {items[1]}"
    return ", ".join(items[:-1]) + f", and {items[-1]}"


def build_theme_system_prompt(labels: list[str], existing_themes: list[str] = None) -> str:
    n = len(labels)
    annotated_labels = [f"{labels[0]} (oldest)"] + labels[1:-1] + [f"{labels[-1]} (most recent)"]
    label_description = _oxford_join(annotated_labels)
    delta_list = ", ".join(f"{labels[i]} vs {labels[i + 1]}" for i in range(n - 1))
    json_skeleton = "{\n" + ",\n".join(
        f'  "<period of {label}>": {{"themes": ["...", "...", "..."]}}' for label in labels[1:]
    ) + "\n}"

    step5 = ""
    existing_section = ""
    if existing_themes:
        step5 = "\n5. HISTORICAL THEME RECONCILIATION. After completing step 4, compare each selected theme name against the EXISTING THEMES list at the end of this prompt. If a selected theme is semantically the same underlying topic as an existing theme, rename it to match that existing name exactly (preserve its exact spelling and capitalization). Apply this only when the match is strong — the two names clearly describe the same investment topic. If no existing theme is a close semantic match, keep the name from step 4. Do not force a weak or partial match."
        existing_section = "\n\nEXISTING THEMES FOR RECONCILIATION (step 5):\n" + "\n".join(f"- {t}" for t in existing_themes)

    return f"""You are given {n} quarters of earnings call Q&A bullet-point summaries for the same company, ordered chronologically and labeled {label_description}.

For each consecutive delta ({delta_list}):
1. Find 3 common themes that are heavily discussed during the Q&A sessions of both quarters in that delta.
2. Find 2 new or emerging themes that are heavily discussed during the Q&A session of the later quarter in that delta but were lightly or never discussed during the earlier quarter.
3. Rank all 5 of those themes (the 3 common plus the 2 emerging) by how much discussion volume each received in the LATER quarter's Q&A session only (ignore the earlier quarter's volume for this ranking). Keep only the top 3 themes by that ranking, ordered from highest to lowest volume.
4. CROSS-DELTA CONSOLIDATION. After the top-3 themes have been selected for every delta, compare theme names across deltas (not within the same delta — that's already handled below). If a theme chosen for one delta is semantically the same underlying topic as a theme chosen for another delta, rename both occurrences to one common, consistent name (2-3 words, high-level and neutral, per the Rules below) so the same underlying theme is tracked identically wherever it recurs. Do not change the ranking, the order, or which themes were selected — only normalize the name.{step5}

Rules:
- "Heavily discussed" / theme volume is judged by the amount of Q&A dialogue (number of words) spent discussing the theme.
- Each theme must be very high-level and neutral (e.g. "Proprietary Silicon", not "Proprietary Silicon Development Timing").
- Each theme in a delta quarter must be over 80% different semantically compared to the others.
- Each theme name must be 2-3 words.

Output format - return ONLY valid JSON, no commentary, no markdown code fences, no preamble. Structure:
{json_skeleton}
Each delta's key in the JSON must be the period label (e.g. "Q1 2026") of the LATER quarter in that delta. Each delta's "themes" list must contain exactly 3 entries: the top 3 themes after the step-3 ranking and cut, ordered from highest to lowest discussion volume in the later quarter, with no indication of which were originally common vs. emerging.{existing_section}"""


READTHROUGH_DESCRIPTION_LENGTH_RULE = (
    "The two fields combined must be 350-400 characters long, including spaces and punctuation"
)

THEME_DELTA_PROMPT_TEMPLATE = """You are an earnings-call analyst comparing management messaging and analyst tone across a list of investment themes, one quarter-over-quarter delta at a time.

THEME LIST:
{theme_list}

QUARTERS TO ANALYZE:
{quarter_delta}
- DELTA DIRECTION: the quarter tagged (LATER) is the later quarter; (EARLIER) is the prior. Every signal describes how the later quarter changed versus the prior.

RULES (apply to every theme and every step):
- THEME INDEPENDENCE: each theme is analyzed in complete isolation. Evidence, tone, framing, or sentiment from one theme MUST NOT be transferred to another. Management and analyst sides are judged independently per theme; they need not move in the same direction.
- DATA BOUNDARY: use ONLY the two quarters' transcript data provided. Do NOT use outside information.
- ANALYST TONE: the stance reflected in analysts' Q&A questions on a theme — whether they probe risk, press skeptically, or signal confidence, and how management's answers are received.

TASK:
Step 1 - Summary: For each theme in THEME LIST, summarize in a single sentence (1) the management message and (2) the analyst tone in the later quarter.
Step 2 - Signal: For each theme, compare the later quarter against the prior quarter and assign a signal — improved, worsened, or stable.
- If the theme is discussed in the later quarter but not the prior, judge by how positive or negative the later-quarter message and tone are in absolute terms.
- Otherwise judge only by the change versus the prior quarter, not by absolute positivity or negativity. An improved or worsened signal requires a specific, citable QoQ change; if none can be named, the signal is stable. When evidence is ambiguous, default to stable.
- Map each signal to its arrow: improved -> "↑:", worsened -> "↓:", stable -> "→:".
Step 3 - Context: For each theme, summarize in a single sentence why the management message and analyst tone improved, worsened, or stayed stable, citing the QoQ change behind the signal.

OUTPUT CONSTRAINTS:
- {description_length_rule} across the "message"/"tone" field and its "rationale" field combined.
- STYLE: abridged, telegraphic style with abbreviations.
- Omit the final period "." at the end of every field.
- Before producing JSON, verify that: (a) every theme in THEME LIST is present; (b) each rationale's direction matches its signal arrow; (c) every ↑ or ↓ rests on a QoQ change in direction — management or analysts actually moved, not merely repeated a prior stance or posted an in-line/as-guided number. A reaffirmation, an in-line metric, or roughly-offsetting shifts are → (stable). Exception: a theme absent in the prior quarter follows Step 2's absolute-tone rule and is not demoted for lacking a QoQ change.

JSON FORMAT: Output ONLY a valid JSON block using ```json tags, following the exact schema below, with one object per theme in THEME LIST order. Replace the theme placeholders with the exact theme names from THEME LIST. Do not output any text before or after the JSON block. Escape any double quotes inside field values.

```json
{json_format}
```"""


def build_delta_pairs(labels: list[str]) -> list[tuple[str, str]]:
    return [(labels[i + 1], labels[i]) for i in range(len(labels) - 1)]


def build_readthrough_quarter_mapping(ordered_quarters) -> str:
    lines = [f"{item['label']} = {item['period']}" for item in ordered_quarters]
    return "QUARTER MAPPING:\n" + "\n".join(lines)


MERGED_PROMPT_TEMPLATE = """You are an equity research analyst producing three signals per quarter-over-quarter delta from one company's consecutive earnings calls: a sector READ-THROUGH, a BULL forward case, and a BEAR forward case. Inputs are consecutive quarterly transcripts (or telegraphic summaries) ordered oldest to newest; the task is identical regardless of company, sector, fiscal calendar, or quarter count.

QUARTER MAPPING:
{quarter_mapping}
- DELTA: each adjacent pair is one delta; N quarters give N-1 deltas. In a "LATER_vs_EARLIER" key the first quarter is the later one, the second the earlier. Every signal describes how the later quarter changed versus the earlier.

DEFINITIONS:
- EXTERNAL BUSINESS FACTORS: industry/macro conditions management describes that read across to sector peers — end-market demand, customer spending, pricing, input/supply availability, competitive dynamics, regulatory/macro. EXCLUDES the company's own KPIs and internal initiatives (its revenue, margins, segments, roadmap, capital returns, restructuring). The READ-THROUGH uses external factors ONLY.
- FORWARD EXPECTATION: what a bull (or bear) investor would, on that quarter's call, expect for the business over the next 6–9 months. Infer from the whole call but prioritize the Q&A, where management is tested. Bull and bear are assessed separately and need not move in opposite directions.

RULES (apply to every delta and every signal):
- ANALYSIS INDEPENDENCE: read-through, bull, and bear are three independent reads of the same delta. Derive each from its own evidence; never soften, strengthen, or flip one because another track moved. They may point in different directions; sharing an underlying fact is fine as long as each is weighed on its own.
- DATA BOUNDARY: use ONLY the provided transcripts. No outside information.

TASK — for each delta, assign three signals (read-through, bull, bear), each Strengthened / Weakened / Stable:
- READ-THROUGH: select the single dominant external business factor. It MUST be discussed in both quarters of the delta — a factor absent from the earlier quarter is ineligible (no QoQ baseline to measure against). Among eligible factors, pick the one most relevant to sector peers, breaking ties by emphasis (airtime / strongest language). If no external factor appears in both quarters, the read-through is Stable. Then judge how the selected factor's read-through changed QoQ.
- BULL / BEAR: for each side, judge how that side's forward 6–9 month case changed QoQ, prioritizing Q&A evidence.

Signal rule (all three):
- Strengthened: a clear positive QoQ change — escalated language, a raised/affirmed external indicator or guidance, a broadened demand base, a resolved concern (bull/bear: analysts shift from probing a risk to confirming it, corroborated by management).
- Weakened: a clear negative QoQ change — softened/walked-back language, a new or amplified headwind, a narrowed demand base, a new risk management cannot fully address.
- Stable: no clear directional change, or offsetting shifts.
A Strengthened or Weakened signal requires a specific, citable QoQ change in direction — not a reaffirmation, an in-line/as-guided number, or a repeated stance. If none can be named, the signal is Stable. When evidence is ambiguous, default to Stable.
Map each signal to its arrow: Strengthened -> "↑:", Weakened -> "↓:", Stable -> "→:". The read-through arrow fills the "signal" field; each side's arrow fills its "topic" field.

OUTPUT FIELDS (per delta):
- READ-THROUGH "read_through": one sentence stating the later quarter's macro/industry read-through (external factors only). "rationale": one sentence citing one specific QoQ contrast and why it strengthened/weakened/held. {description_length_rule} across the two combined.
- BULL/BEAR "expectation" (per side): one sentence stating the later quarter's forward 6–9 month case. "context": one sentence citing one specific QoQ change in that side's case and why it strengthened/weakened/held. {description_length_rule} across the two combined.
- STYLE: abridged, telegraphic, abbreviations. Omit the final period "." at the end of every field.

Before producing JSON, verify for each delta that: (a) every rationale/context direction matches its arrow; (b) each ↑ or ↓ rests on a real directional QoQ change — else → (stable); (c) the read-through factor was discussed in both quarters; (d) no track's signal was adjusted to agree or disagree with another track's on the same delta.

JSON FORMAT: Output ONLY a valid JSON block using ```json tags, following the exact schema below, with one object per delta matching the periods and order in the schema. No text before or after the JSON block. Escape any double quotes inside field values.

```json
{json_format}
```"""


def build_merged_json_skeleton(labels: list[str]) -> str:
    entries = [
        f"""    {{
      "period": "{later}_vs_{earlier}",
      "signal": "↑: or ↓: or →:",
      "read_through": "one-sentence macro/industry read-through",
      "rationale": "one-sentence QoQ contrast explaining the read-through signal direction",
      "bull": {{
        "topic": "↑: or ↓: or →:",
        "expectation": "one-sentence bullish forward-looking expectation",
        "context": "one-sentence QoQ context explaining the bull signal direction"
      }},
      "bear": {{
        "topic": "↑: or ↓: or →:",
        "expectation": "one-sentence bearish forward-looking expectation",
        "context": "one-sentence QoQ context explaining the bear signal direction"
      }}
    }}"""
        for later, earlier in build_delta_pairs(labels)
    ]
    return "{\n  \"historical_analysis\": [\n" + ",\n".join(entries) + "\n  ]\n}"


def build_merged_system_prompt(ordered_quarters) -> str:
    labels = [item["label"] for item in ordered_quarters]
    return MERGED_PROMPT_TEMPLATE.format(
        quarter_mapping=build_readthrough_quarter_mapping(ordered_quarters),
        description_length_rule=READTHROUGH_DESCRIPTION_LENGTH_RULE,
        json_format=build_merged_json_skeleton(labels),
    )


def build_theme_list(themes: list[str]) -> str:
    return "\n".join(f"- {theme}" for theme in themes)


def build_quarter_delta_mapping(later_item, earlier_item) -> str:
    return (
        f"{later_item['label']} = {later_item['period']} (LATER)\n"
        f"{earlier_item['label']} = {earlier_item['period']} (EARLIER)"
    )


def build_theme_delta_json_skeleton(themes: list[str]) -> str:
    entries = [
        f"""    {{
      "theme": "{theme}",
      "mgmt": {{
        "signal": "↑: or ↓: or →:",
        "message": "one-sentence management message",
        "rationale": "one-sentence QoQ contrast explaining the signal direction"
      }},
      "analyst": {{
        "signal": "↑: or ↓: or →:",
        "tone": "one-sentence analyst tone",
        "rationale": "one-sentence QoQ contrast explaining the signal direction"
      }}
    }}"""
        for theme in themes
    ]
    return "{\n  \"theme_analysis\": [\n" + ",\n".join(entries) + "\n  ]\n}"


def build_theme_delta_system_prompt(themes: list[str], later_item, earlier_item) -> str:
    return THEME_DELTA_PROMPT_TEMPLATE.format(
        theme_list=build_theme_list(themes),
        quarter_delta=build_quarter_delta_mapping(later_item, earlier_item),
        description_length_rule=READTHROUGH_DESCRIPTION_LENGTH_RULE,
        json_format=build_theme_delta_json_skeleton(themes),
    )


# ============================================================
# ANALYSIS CALLS
# ============================================================

def extract_qa_section(content: str) -> str:
    match = re.search(r"QUESTION AND ANSWER SECTION", content, flags=re.IGNORECASE)
    return content[match.start():] if match else content


def build_themes_params(ordered_quarters, existing_themes: list[str] = None) -> dict:
    header = "Input Periods: " + ", ".join(
        f"{item['label']}:{item['period']}" for item in ordered_quarters
    )
    sections = "\n\n".join(
        f"=== {item['label']} ({item['period']}) ===\n{extract_qa_section(item['content'])}"
        for item in ordered_quarters
    )
    user_content = f"{header}\n\n{sections}"
    system_prompt = build_theme_system_prompt([item["label"] for item in ordered_quarters], existing_themes)
    return {
        "model": MODEL,
        "max_tokens": 64000,
        "thinking": {"type": "adaptive"},
        "output_config": {"effort": THEMES_EFFORT},
        "system": system_prompt,
        "messages": [{"role": "user", "content": user_content}],
    }


def parse_themes_response(raw_text: str):
    cleaned = re.sub(r"^```(?:json)?|```$", "", raw_text.strip(), flags=re.MULTILINE).strip()
    try:
        return json.loads(cleaned), None
    except json.JSONDecodeError as e:
        return None, f"Failed to parse JSON response: {e}\n\nRaw response:\n{raw_text}"


def parse_fenced_json_response(raw_text: str):
    match = re.search(r"```json\s*(.*?)```", raw_text, flags=re.DOTALL)
    if not match:
        return None, f"Failed to find JSON block in response\n\nRaw response:\n{raw_text}"
    try:
        return json.loads(match.group(1).strip()), None
    except json.JSONDecodeError as e:
        return None, f"Failed to parse JSON response: {e}\n\nRaw response:\n{raw_text}"


def build_merged_params(ordered_quarters) -> dict:
    sections = "\n\n".join(
        f"=== {item['label']} ({item['period']}) ===\n{item['content']}"
        for item in ordered_quarters
    )
    return {
        "model": READTHROUGH_MODEL,
        "max_tokens": 64000,
        "thinking": {"type": "adaptive"},
        "output_config": {"effort": READTHROUGH_EFFORT},
        "system": build_merged_system_prompt(ordered_quarters),
        "messages": [{"role": "user", "content": sections}],
    }


def build_theme_delta_params(themes: list[str], later_item, earlier_item) -> dict:
    sections = "\n\n".join(
        f"=== {item['label']} ({item['period']}) ===\n{item['content']}"
        for item in (later_item, earlier_item)
    )
    return {
        "model": THEME_DELTA_MODEL,
        "max_tokens": 64000,
        "thinking": {"type": "adaptive"},
        "output_config": {"effort": THEME_DELTA_EFFORT},
        "system": build_theme_delta_system_prompt(themes, later_item, earlier_item),
        "messages": [{"role": "user", "content": sections}],
    }


def build_themes_batch_request(custom_id: str, ordered_quarters, existing_themes: list[str] = None) -> dict:
    return {"custom_id": custom_id, "params": build_themes_params(ordered_quarters, existing_themes)}


def build_merged_batch_request(custom_id: str, ordered_quarters) -> dict:
    return {"custom_id": custom_id, "params": build_merged_params(ordered_quarters)}


def build_theme_delta_batch_request(custom_id: str, themes: list[str], later_item, earlier_item) -> dict:
    return {"custom_id": custom_id, "params": build_theme_delta_params(themes, later_item, earlier_item)}


def _stream_text(client: Anthropic, params: dict) -> str:
    with client.messages.stream(**params) as stream:
        response = stream.get_final_message()
    return "".join(block.text for block in response.content if block.type == "text")


def analyze_themes(client: Anthropic, ordered_quarters, existing_themes: list[str] = None):
    raw_text = _stream_text(client, build_themes_params(ordered_quarters, existing_themes))
    return parse_themes_response(raw_text)


def analyze_merged(client: Anthropic, ordered_quarters):
    raw_text = _stream_text(client, build_merged_params(ordered_quarters))
    return parse_fenced_json_response(raw_text)


def analyze_theme_delta(client: Anthropic, themes: list[str], later_item, earlier_item):
    raw_text = _stream_text(client, build_theme_delta_params(themes, later_item, earlier_item))
    return parse_fenced_json_response(raw_text)


# ============================================================
# BATCH PROCESSING
# ============================================================

def submit_stage1_batch(client: Anthropic, ordered_quarters, existing_themes: list[str] = None) -> str:
    requests = [
        build_themes_batch_request("themes", ordered_quarters, existing_themes),
        build_merged_batch_request("merged", ordered_quarters),
    ]
    batch = client.messages.batches.create(requests=requests)
    return batch.id


def submit_stage2_batch(client: Anthropic, themes_result: dict, ordered_quarters):
    requests = []
    manifest = {}
    for later_item, earlier_item in zip(ordered_quarters[1:], ordered_quarters[:-1]):
        delta_key = f"{later_item['label']}_vs_{earlier_item['label']}"
        ranked_themes = themes_result.get(later_item["period"], {}).get("themes", [])
        if not ranked_themes:
            continue
        merged_id = f"{delta_key}__themes"
        merged_themes = COMMON_THEME_SET + ranked_themes
        requests.append(build_theme_delta_batch_request(merged_id, merged_themes, later_item, earlier_item))
        manifest[delta_key] = {
            "merged_id": merged_id,
            "ranked_themes": ranked_themes,
        }
    batch = client.messages.batches.create(requests=requests)
    return batch.id, manifest


def split_merged_result(merged_result: dict):
    entries = merged_result.get("historical_analysis", []) if merged_result else []
    readthrough_result = {
        "historical_analysis": [
            {
                "period": e["period"],
                "signal": e["signal"],
                "read_through": e["read_through"],
                "rationale": e["rationale"],
            }
            for e in entries
        ]
    }
    bullbear_result = {
        "historical_analysis": [
            {"period": e["period"], "bull": e["bull"], "bear": e["bear"]}
            for e in entries
        ]
    }
    return readthrough_result, bullbear_result


def split_theme_delta_result(merged_result: dict, ranked_themes: list):
    ranked_set = set(ranked_themes)
    entries = merged_result.get("theme_analysis", []) if merged_result else []
    common_result = {"theme_analysis": [e for e in entries if e["theme"] not in ranked_set]}
    ranked_result = {"theme_analysis": [e for e in entries if e["theme"] in ranked_set]}
    return common_result, ranked_result


def fetch_raw_batch_results(client: Anthropic, batch_id: str) -> dict:
    results = {}
    for entry in client.messages.batches.results(batch_id):
        if entry.result.type == "succeeded":
            content = entry.result.message.content
            text = "".join(block.text for block in content if block.type == "text")
            block_summary = [{"type": block.type} for block in content]
            results[entry.custom_id] = (text, None, block_summary)
        else:
            error_message = getattr(entry.result, "error", None)
            error_str = str(error_message) if error_message is not None else entry.result.type
            results[entry.custom_id] = (None, error_str, None)
    return results


# ============================================================
# UTILITIES
# ============================================================

def parse_md_metadata(filename: str):
    name = filename.rsplit(".", 1)[0]
    parts = name.split("-")
    ticker = " ".join(parts[:2]) if len(parts) >= 2 else None
    period_match = re.search(r"Q[1-4]\s+20\d{2}", name)
    period = period_match.group(0) if period_match else None
    return ticker, period


def quarter_sort_key(period: str):
    quarter, year = period.split()
    return int(year), int(quarter[1])


# ============================================================
# STREAMLIT UI
# ============================================================

st.set_page_config(page_title="Cross-Quarter Theme Analysis", layout="wide")
st.title("Cross-Quarter Theme Analysis")

with st.sidebar:
    api_key = st.text_input("Anthropic API key", type="password")
    db_url = st.text_input(
        "Supabase connection string",
        type="password",
        help="postgresql://postgres:PASSWORD@db.REF.supabase.co:5432/postgres  "
             "(Settings → Database → Connection string → Direct)",
    )

theme_md_files = st.file_uploader(
    "Upload telegraphic .md files (one or more companies, 2 or more consecutive quarters each)",
    type=["md"],
    accept_multiple_files=True,
)

db_file = st.file_uploader(
    "Upload Database.xlsx",
    type=["xlsx"],
)


def run_batch_analysis_for_group(client: Anthropic, group: dict, conn, status):
    parsed_quarters = group["quarters"]
    existing_themes = group["existing_themes"]
    ticker = group["ticker"]
    quarter_id_map = group["quarter_id_map"]

    status.info("Submitting stage 1 batch (themes, read-through, bull/bear)...")
    batch_id = submit_stage1_batch(client, parsed_quarters, existing_themes)

    while True:
        batch = client.messages.batches.retrieve(batch_id)
        if batch.processing_status == "ended":
            break
        counts = batch.request_counts
        status.info(
            f"Stage 1 running — processing: {counts.processing}, "
            f"succeeded: {counts.succeeded}, errored: {counts.errored}. Checking again in 30s..."
        )
        time.sleep(30)

    status.info("Fetching stage 1 results...")
    raw_results = fetch_raw_batch_results(client, batch_id)

    themes_text, themes_err, _ = raw_results.get("themes", (None, "missing", None))
    themes_result = None
    if themes_text:
        themes_result, parse_err = parse_themes_response(themes_text)
        if parse_err:
            st.error(f"themes: {parse_err}")
            with st.expander("themes raw response"):
                st.text(themes_text)
    else:
        st.error(f"themes batch error: {themes_err}")

    readthrough_result = None
    bullbear_result = None
    merged_text, merged_err, merged_blocks = raw_results.get("merged", (None, "missing", None))
    if merged_text is not None:
        if not merged_text:
            st.error(f"merged: batch succeeded but response contained no text (blocks: {merged_blocks})")
        else:
            merged_result, merged_parse_err = parse_fenced_json_response(merged_text)
            if merged_parse_err:
                st.error(f"merged: JSON parse failed — {merged_parse_err}")
                with st.expander("merged raw response"):
                    st.text(merged_text)
            else:
                readthrough_result, bullbear_result = split_merged_result(merged_result)
    else:
        st.error(f"merged batch error: {merged_err}")

    if themes_result is None:
        status.error("Stage 1 failed to produce themes — cannot continue.")
        return

    st.json(themes_result)

    status.info("Submitting stage 2 batch (per-delta theme analysis)...")
    stage2_id, manifest = submit_stage2_batch(client, themes_result, parsed_quarters)

    while True:
        batch = client.messages.batches.retrieve(stage2_id)
        if batch.processing_status == "ended":
            break
        counts = batch.request_counts
        status.info(
            f"Stage 2 running — processing: {counts.processing}, "
            f"succeeded: {counts.succeeded}, errored: {counts.errored}. Checking again in 30s..."
        )
        time.sleep(30)

    status.info("Fetching stage 2 results and writing to Supabase...")
    raw_results = fetch_raw_batch_results(client, stage2_id)

    theme_delta_by_key = {}
    for delta_key, ids in manifest.items():
        merged_text, merged_err, merged_blocks = raw_results.get(ids["merged_id"], (None, "missing", None))
        merged_result = None
        if merged_text is not None:
            if not merged_text:
                st.error(f"{delta_key}: batch succeeded but response contained no text (blocks: {merged_blocks})")
            else:
                merged_result, merged_parse_err = parse_fenced_json_response(merged_text)
                if merged_parse_err:
                    st.error(f"{delta_key}: JSON parse failed — {merged_parse_err}")
                    with st.expander(f"{delta_key} raw response"):
                        st.text(merged_text)
        else:
            st.error(f"{delta_key} batch error: {merged_err}")

        common_result, ranked_result = split_theme_delta_result(merged_result, ids["ranked_themes"])
        theme_delta_by_key[delta_key] = {
            "common": common_result,
            "ranked": ranked_result,
            "ranked_themes": ids["ranked_themes"],
        }

    write_analysis_results(
        conn, ticker, quarter_id_map, parsed_quarters,
        readthrough_result, bullbear_result, theme_delta_by_key,
    )

    st.json(readthrough_result)
    st.json(bullbear_result)
    st.json(theme_delta_by_key)
    status.success(f"{ticker}: results written to Supabase.")


def run_interactive_analysis_for_group(client: Anthropic, group: dict, conn):
    parsed_quarters = group["quarters"]
    existing_themes = group["existing_themes"]
    ticker = group["ticker"]
    quarter_id_map = group["quarter_id_map"]

    with st.spinner(f"{ticker}: analyzing cross-quarter themes..."):
        themes, error = analyze_themes(client, parsed_quarters, existing_themes)
    if error:
        st.error(error)
        return

    st.json(themes)

    with st.spinner(f"{ticker}: analyzing sector read-through and bull/bear signals..."):
        merged, merged_error = analyze_merged(client, parsed_quarters)
    readthrough_result = None
    bullbear_result = None
    if merged_error:
        st.error(merged_error)
    else:
        readthrough_result, bullbear_result = split_merged_result(merged)
        st.json(readthrough_result)
        st.json(bullbear_result)

    theme_delta_by_key = {}
    for later_item, earlier_item in zip(parsed_quarters[1:], parsed_quarters[:-1]):
        delta_key = f"{later_item['label']}_vs_{earlier_item['label']}"
        ranked_themes = themes.get(later_item["period"], {}).get("themes", [])
        if not ranked_themes:
            st.error(f"{delta_key}: no ranked themes found for {later_item['period']}")
            continue

        merged_themes = COMMON_THEME_SET + ranked_themes
        with st.spinner(f"{ticker}: analyzing themes for {delta_key}..."):
            merged_result, merged_error = analyze_theme_delta(client, merged_themes, later_item, earlier_item)

        if merged_error:
            st.error(f"{delta_key}: {merged_error}")

        common_result, ranked_result = split_theme_delta_result(merged_result, ranked_themes)

        theme_delta_by_key[delta_key] = {
            "common": common_result,
            "ranked": ranked_result,
            "ranked_themes": ranked_themes,
        }
        st.write(f"Theme delta analysis — {delta_key}")
        st.json(theme_delta_by_key[delta_key])

    write_analysis_results(
        conn, ticker, quarter_id_map, parsed_quarters,
        readthrough_result, bullbear_result, theme_delta_by_key,
    )


if theme_md_files:
    parsed_files = []
    for f in theme_md_files:
        ticker, period = parse_md_metadata(f.name)
        if period is None:
            st.error(f"{f.name}: could not parse a quarter/period (e.g. 'Q1 2026') from filename")
            continue
        parsed_files.append(
            {
                "filename": f.name,
                "ticker": ticker,
                "period": period,
                "content": f.read().decode("utf-8"),
            }
        )

    if not parsed_files:
        st.warning(f"Need at least {MIN_QUARTERS} valid .md files; found 0.")
    else:
        files_by_ticker = {}
        for item in parsed_files:
            files_by_ticker.setdefault(item["ticker"], []).append(item)

        workbook = openpyxl.load_workbook(db_file, data_only=True) if db_file else None
        conn = None
        quarter_id_map = {}

        if workbook is not None and db_url:
            with st.spinner("Syncing Excel structure to Supabase..."):
                try:
                    conn = get_connection(db_url)
                    quarter_id_map = sync_workbook_structure(conn, workbook)
                    st.success(f"Synced {len(quarter_id_map)} tickers to Supabase.")
                except Exception as e:
                    st.error(f"Supabase connection failed: {e}")

        cur = conn.cursor() if conn else None
        ticker_groups = []

        for ticker, items in files_by_ticker.items():
            if len(items) < MIN_QUARTERS:
                st.warning(
                    f"{ticker}: need at least {MIN_QUARTERS} valid .md files; found {len(items)}. Skipping."
                )
                continue

            items.sort(key=lambda item: quarter_sort_key(item["period"]))
            labels = build_quarter_labels(len(items))
            for label, item in zip(labels, items):
                item["label"] = label

            group = {
                "ticker": ticker,
                "quarters": items,
                "quarter_id_map": quarter_id_map.get(ticker, {}),
                "existing_themes": [],
                "db_ready": False,
            }

            with st.expander(f"{ticker} — {len(items)} quarters", expanded=True):
                st.write({item["label"]: f"{item['ticker']} {item['period']}" for item in items})

                if workbook is None:
                    st.info(f"Upload Database.xlsx above to sync '{ticker}' to Supabase.")
                elif not db_url:
                    st.info("Enter your Supabase connection string in the sidebar to continue.")
                elif conn is None:
                    st.error("Supabase connection failed — check your connection string.")
                elif ticker not in quarter_id_map:
                    st.error(f"No tab found matching ticker '{ticker}' in the uploaded Excel.")
                else:
                    missing_periods = [
                        item["period"] for item in items
                        if period_to_quarter(item["period"]) not in quarter_id_map[ticker]
                    ]
                    if missing_periods:
                        st.error(
                            f"No row found for: {', '.join(missing_periods)} "
                            f"(expected column A format like '{period_to_quarter(items[0]['period'])}')"
                        )
                    else:
                        st.success(
                            "Matched rows: "
                            + ", ".join(
                                f"{item['period']} -> id {quarter_id_map[ticker][period_to_quarter(item['period'])]}"
                                for item in items
                            )
                        )
                        group["existing_themes"] = get_existing_variable_themes(cur, ticker)
                        group["db_ready"] = True

            ticker_groups.append(group)

        if cur:
            cur.close()

        ready_groups = [g for g in ticker_groups if g["db_ready"]]

        use_batch = st.checkbox(
            "Use batch processing (cheaper, ~50% cost)",
            help="Submits the analysis as Batches API jobs at ~50% lower API cost. "
            "Runs automatically: stage 1 (themes, read-through, bull/bear) runs first, "
            "then stage 2 (per-delta theme analysis) is submitted and awaited. "
            "Tickers are processed one at a time. Keep this tab open until complete.",
        )

        if use_batch:
            if st.button("Run batch", disabled=not (api_key and ready_groups)):
                client = Anthropic(api_key=api_key)
                for group in ready_groups:
                    st.subheader(group["ticker"])
                    status = st.empty()
                    run_batch_analysis_for_group(client, group, conn, status)
            elif not api_key:
                st.info("Enter your Anthropic API key in the sidebar to begin.")

        elif st.button("Analyze Themes", disabled=not (api_key and ready_groups)):
            client = Anthropic(api_key=api_key)
            for group in ready_groups:
                st.subheader(group["ticker"])
                run_interactive_analysis_for_group(client, group, conn)

elif not api_key:
    st.info("Enter your Anthropic API key in the sidebar to begin.")
else:
    st.info(f"Upload at least {MIN_QUARTERS} telegraphic .md files to begin.")
