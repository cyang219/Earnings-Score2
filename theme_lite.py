import io
import json
import re
import time

import openpyxl
import streamlit as st
from anthropic import Anthropic
from openpyxl.utils import column_index_from_string

MODEL = "claude-sonnet-4-6"
THEMES_EFFORT = "medium"
THEME_DELTA_EFFORT = "medium"

MIN_QUARTERS = 2

DESCRIPTION_LENGTH_RULE = (
    "The two fields combined must be 350-400 characters long, including spaces and punctuation"
)


# ============================================================
# PROMPTS
# ============================================================

def build_quarter_labels(n: int) -> list[str]:
    return [f"Q-{n - 1 - i}" for i in range(n - 1)] + ["Latest_Q"]


def _oxford_join(items: list[str]) -> str:
    if len(items) == 1:
        return items[0]
    if len(items) == 2:
        return f"{items[0]} and {items[1]}"
    return ", ".join(items[:-1]) + f", and {items[-1]}"


def build_theme_system_prompt(labels: list[str], existing_themes: list[str] = None) -> str:
    n = len(labels)
    annotated_labels = [f"{labels[0]} (oldest)"] + labels[1:-1] + [f"{labels[-1]} (most recent)"]
    label_description = _oxford_join(annotated_labels)
    delta_list = ", ".join(f"{labels[i]} vs {labels[i + 1]}" for i in range(n - 1))
    json_skeleton = "{\n" + ",\n".join(
        f'  "<period of {label}>": {{"themes": ["...", "...", "..."]}}' for label in labels[1:]
    ) + "\n}"

    step5 = ""
    existing_section = ""
    if existing_themes:
        step5 = "\n5. HISTORICAL THEME RECONCILIATION. After completing step 4, compare each selected theme name against the EXISTING THEMES list at the end of this prompt. If a selected theme is semantically the same underlying topic as an existing theme, rename it to match that existing name exactly (preserve its exact spelling and capitalization). Apply this only when the match is strong — the two names clearly describe the same investment topic. If no existing theme is a close semantic match, keep the name from step 4. Do not force a weak or partial match."
        existing_section = "\n\nEXISTING THEMES FOR RECONCILIATION (step 5):\n" + "\n".join(f"- {t}" for t in existing_themes)

    return f"""You are given {n} quarters of earnings call Q&A bullet-point summaries for the same company, ordered chronologically and labeled {label_description}.

For each consecutive delta ({delta_list}):
1. Find 3 common themes that are heavily discussed during the Q&A sessions of both quarters in that delta.
2. Find 2 new or emerging themes that are heavily discussed during the Q&A session of the later quarter in that delta but were lightly or never discussed during the earlier quarter.
3. Rank all 5 of those themes (the 3 common plus the 2 emerging) by how much discussion volume each received in the LATER quarter's Q&A session only (ignore the earlier quarter's volume for this ranking). Keep only the top 3 themes by that ranking, ordered from highest to lowest volume.
4. CROSS-DELTA CONSOLIDATION. After the top-3 themes have been selected for every delta, compare theme names across deltas (not within the same delta — that's already handled below). If a theme chosen for one delta is semantically the same underlying topic as a theme chosen for another delta, rename both occurrences to one common, consistent name (2-3 words, high-level and neutral, per the Rules below) so the same underlying theme is tracked identically wherever it recurs. Do not change the ranking, the order, or which themes were selected — only normalize the name.{step5}

Rules:
- "Heavily discussed" / theme volume is judged by the amount of Q&A dialogue (number of words) spent discussing the theme.
- Each theme must be very high-level and neutral (e.g. "Proprietary Silicon", not "Proprietary Silicon Development Timing").
- Each theme in a delta quarter must be over 80% different semantically compared to the others.
- Each theme name must be 2-3 words.

Output format - return ONLY valid JSON, no commentary, no markdown code fences, no preamble. Structure:
{json_skeleton}
Each delta's key in the JSON must be the period label (e.g. "Q1 2026") of the LATER quarter in that delta. Each delta's "themes" list must contain exactly 3 entries: the top 3 themes after the step-3 ranking and cut, ordered from highest to lowest discussion volume in the later quarter, with no indication of which were originally common vs. emerging.{existing_section}"""


THEME_DELTA_PROMPT_TEMPLATE = """You are an earnings-call analyst comparing management messaging and analyst tone across investment themes for one quarter-over-quarter delta. You will be given consecutive telegraphic summaries of quarterly earnings calls from a single publicly traded company.

DEFINITIONS (canonical — all later references point back here):
- MANAGEMENT MESSAGE = what management conveys about a theme in the prepared remarks and their Q&A answers: the substance, confidence, and framing of their commentary on that theme.
- ANALYST TONE = the stance reflected in analysts' Q&A questions on a theme: whether they probe risk, press skeptically, or signal confidence, and how management's answers are received.
- DELTA DIRECTION = the FIRST-named quarter in QUARTERS TO ANALYZE is the LATER quarter; the SECOND-named is the EARLIER quarter. A signal describes how the later quarter changed versus the earlier quarter.
- THEME INDEPENDENCE = each theme is analyzed in complete isolation. Evidence, tone, framing, or sentiment from one theme MUST NOT be transferred to another. Management and analyst sides are also judged independently per theme; they need not move in the same direction.
- DATA BOUNDARY = use ONLY the two quarters' transcript data provided. Do NOT use outside information.

THEME LIST:
{theme_list}

QUARTERS TO ANALYZE:
{quarter_delta}

TASK:

Step 1: PER-THEME READ. For each theme in THEME LIST, using only the two quarters (see DATA BOUNDARY), internally summarize the later quarter's management message and analyst tone for that theme, and how each compares to the earlier quarter. If a theme is absent in the earlier quarter, note this now so the special-case rule in Step 2 applies correctly. Process each theme fully and in isolation before moving to the next (see THEME INDEPENDENCE). Use this only as internal reasoning; do not output Step 1.

Step 2: SIGNAL SCRATCHPAD. Start a <scratchpad> assigning a management signal and an analyst signal to each theme. Judge each signal only by the change versus the earlier quarter (see DELTA DIRECTION), not by absolute positivity or negativity.

Classify each side (management, then analyst) for each theme using this rule, then assign the signal:
- Improved: a clear positive change vs the earlier quarter — management escalates confidence/affirms progress/resolves a prior concern on that theme, OR analyst questions shift from probing a risk toward confirming an improvement.
- Worsened: a clear negative change — management softens, walks back, or introduces a new concern on that theme, OR analyst questions surface a new risk management cannot fully address.
- Stable: no clear directional change — framing, confidence, and tone on that theme are materially unchanged, OR positive and negative shifts roughly offset.
An Improved or Worsened signal requires a specific, citable QoQ change on that theme for that side (see DATA BOUNDARY). If no such concrete change can be named, the signal must be Stable. When evidence is genuinely balanced or ambiguous, default to Stable rather than guessing a direction.
Special case — theme absent in the earlier quarter: if the theme is not discussed in the earlier quarter, judge by the later quarter's absolute message/tone: positive = Improved, negative = Worsened, neutral or limited = Stable. Do not mark Improved or Worsened solely because the theme appeared.

Copy the exact template below and fill in the brackets, producing one line per theme in THEME LIST order. Do NOT write anything else inside the scratchpad.
<scratchpad>
{scratchpad_format}
</scratchpad>
- Signal must be exactly one of: Improved, Worsened, Stable
- Each signal must be evidence-linked to that theme's own data (see THEME INDEPENDENCE)

Step 3: SIGNAL FIELD. Derive the "signal" field mechanically from the scratchpad Signal for that same theme and side (not an independent judgment):
- Improved -> "↑:"   |   Worsened -> "↓:"   |   Stable -> "→:"
If the field and the scratchpad Signal would ever disagree, the scratchpad Signal is authoritative.

Step 4: MESSAGE/TONE AND RATIONALE FIELDS. For each theme, produce three fields per side (mgmt and analyst): "signal" (the arrow from Step 3), then for mgmt "message" and for analyst "tone", and "rationale" for both:
- "message" (mgmt) / "tone" (analyst): one sentence stating the later-quarter management message or analyst tone on that theme.
- "rationale": one sentence (a) citing one specific QoQ change on that theme and (b) explaining why it improved, worsened, or stayed stable, matching that side's signal direction.
- {description_length_rule} across the "message"/"tone" field and its "rationale" field combined.
- STYLE: abridged, telegraphic style with abbreviations.
- Omit the final period "." at the end of every field.

Before producing the JSON, verify that (a) every theme in THEME LIST is present, (b) each theme's mgmt and analyst signals comply with THEME INDEPENDENCE, and (c) each rationale's described direction matches its signal arrow and the scratchpad Signal. If any disagree, the scratchpad Signal governs; revise to match.

JSON FORMAT: After </scratchpad>, output ONLY a valid JSON block using ```json tags, following the exact schema below, with one object per theme in THEME LIST order. Replace the theme placeholders with the exact theme names from THEME LIST. Do not output any text between </scratchpad> and the JSON block, or after the JSON block. Escape any double quotes inside field values.

```json
{json_format}
```"""


def build_delta_pairs(labels: list[str]) -> list[tuple[str, str]]:
    return [(labels[i + 1], labels[i]) for i in range(len(labels) - 1)]


def build_theme_list(themes: list[str]) -> str:
    return "\n".join(f"- {theme}" for theme in themes)


def build_quarter_delta_mapping(later_item, earlier_item) -> str:
    return (
        f"{later_item['label']} = {later_item['period']} (LATER)\n"
        f"{earlier_item['label']} = {earlier_item['period']} (EARLIER)"
    )


def build_theme_delta_scratchpad_format(themes: list[str]) -> str:
    return "\n".join(
        f"{theme} -> Mgmt Signal: [Signal]; Mgmt Driver: [2–5 word evidence]; "
        f"Analyst Signal: [Signal]; Analyst Driver: [2–5 word evidence]"
        for theme in themes
    )


def build_theme_delta_json_skeleton(themes: list[str]) -> str:
    entries = [
        f"""    {{
      "theme": "{theme}",
      "mgmt": {{
        "signal": "↑: or ↓: or →:",
        "message": "one-sentence management message",
        "rationale": "one-sentence QoQ contrast explaining the signal direction"
      }},
      "analyst": {{
        "signal": "↑: or ↓: or →:",
        "tone": "one-sentence analyst tone",
        "rationale": "one-sentence QoQ contrast explaining the signal direction"
      }}
    }}"""
        for theme in themes
    ]
    return "{\n  \"theme_analysis\": [\n" + ",\n".join(entries) + "\n  ]\n}"


def build_theme_delta_system_prompt(themes: list[str], later_item, earlier_item) -> str:
    return THEME_DELTA_PROMPT_TEMPLATE.format(
        theme_list=build_theme_list(themes),
        quarter_delta=build_quarter_delta_mapping(later_item, earlier_item),
        scratchpad_format=build_theme_delta_scratchpad_format(themes),
        description_length_rule=DESCRIPTION_LENGTH_RULE,
        json_format=build_theme_delta_json_skeleton(themes),
    )


# ============================================================
# ANALYSIS CALLS
# ============================================================

def extract_qa_section(content: str) -> str:
    match = re.search(r"QUESTION AND ANSWER SECTION", content, flags=re.IGNORECASE)
    return content[match.start():] if match else content


def build_themes_params(ordered_quarters, existing_themes: list[str] = None) -> dict:
    header = "Input Periods: " + ", ".join(
        f"{item['label']}:{item['period']}" for item in ordered_quarters
    )
    sections = "\n\n".join(
        f"=== {item['label']} ({item['period']}) ===\n{extract_qa_section(item['content'])}"
        for item in ordered_quarters
    )
    user_content = f"{header}\n\n{sections}"
    system_prompt = build_theme_system_prompt([item["label"] for item in ordered_quarters], existing_themes)
    return {
        "model": MODEL,
        "max_tokens": 64000,
        "thinking": {"type": "adaptive"},
        "output_config": {"effort": THEMES_EFFORT},
        "system": system_prompt,
        "messages": [{"role": "user", "content": user_content}],
    }


def build_theme_delta_params(themes: list[str], later_item, earlier_item) -> dict:
    sections = "\n\n".join(
        f"=== {item['label']} ({item['period']}) ===\n{item['content']}"
        for item in (later_item, earlier_item)
    )
    return {
        "model": MODEL,
        "max_tokens": 64000,
        "thinking": {"type": "adaptive"},
        "output_config": {"effort": THEME_DELTA_EFFORT},
        "system": build_theme_delta_system_prompt(themes, later_item, earlier_item),
        "messages": [{"role": "user", "content": sections}],
    }


def parse_themes_response(raw_text: str):
    cleaned = re.sub(r"^```(?:json)?|```$", "", raw_text.strip(), flags=re.MULTILINE).strip()
    try:
        return json.loads(cleaned), None
    except json.JSONDecodeError as e:
        return None, f"Failed to parse JSON response: {e}\n\nRaw response:\n{raw_text}"


def parse_fenced_json_response(raw_text: str):
    match = re.search(r"```json\s*(.*?)```", raw_text, flags=re.DOTALL)
    if not match:
        return None, f"Failed to find JSON block in response\n\nRaw response:\n{raw_text}"
    try:
        return json.loads(match.group(1).strip()), None
    except json.JSONDecodeError as e:
        return None, f"Failed to parse JSON response: {e}\n\nRaw response:\n{raw_text}"


def _stream_text(client: Anthropic, params: dict) -> str:
    with client.messages.stream(**params) as stream:
        response = stream.get_final_message()
    return "".join(block.text for block in response.content if block.type == "text")


def analyze_themes(client: Anthropic, ordered_quarters, existing_themes: list[str] = None):
    raw_text = _stream_text(client, build_themes_params(ordered_quarters, existing_themes))
    return parse_themes_response(raw_text)


def analyze_theme_delta(client: Anthropic, themes: list[str], later_item, earlier_item):
    raw_text = _stream_text(client, build_theme_delta_params(themes, later_item, earlier_item))
    return parse_fenced_json_response(raw_text)


# ============================================================
# BATCH PROCESSING
# ============================================================

def build_themes_batch_request(custom_id: str, ordered_quarters, existing_themes: list[str] = None) -> dict:
    return {"custom_id": custom_id, "params": build_themes_params(ordered_quarters, existing_themes)}


def build_theme_delta_batch_request(custom_id: str, themes: list[str], later_item, earlier_item) -> dict:
    return {"custom_id": custom_id, "params": build_theme_delta_params(themes, later_item, earlier_item)}


def submit_stage1_batch(client: Anthropic, ordered_quarters, existing_themes: list[str] = None) -> str:
    requests = [build_themes_batch_request("themes", ordered_quarters, existing_themes)]
    batch = client.messages.batches.create(requests=requests)
    return batch.id


def submit_stage2_batch(client: Anthropic, themes_result: dict, ordered_quarters):
    requests = []
    manifest = {}
    for later_item, earlier_item in zip(ordered_quarters[1:], ordered_quarters[:-1]):
        delta_key = f"{later_item['label']}_vs_{earlier_item['label']}"
        ranked_themes = themes_result.get(later_item["period"], {}).get("themes", [])
        if not ranked_themes:
            continue
        theme_id = f"{delta_key}__themes"
        requests.append(build_theme_delta_batch_request(theme_id, ranked_themes, later_item, earlier_item))
        manifest[delta_key] = {"theme_id": theme_id, "ranked_themes": ranked_themes}
    batch = client.messages.batches.create(requests=requests)
    return batch.id, manifest


def fetch_raw_batch_results(client: Anthropic, batch_id: str) -> dict:
    results = {}
    for entry in client.messages.batches.results(batch_id):
        if entry.result.type == "succeeded":
            content = entry.result.message.content
            text = "".join(block.text for block in content if block.type == "text")
            block_summary = [{"type": block.type} for block in content]
            results[entry.custom_id] = (text, None, block_summary)
        else:
            error_message = getattr(entry.result, "error", None)
            error_str = str(error_message) if error_message is not None else entry.result.type
            results[entry.custom_id] = (None, error_str, None)
    return results


# ============================================================
# DB EXPORT
# ============================================================

def find_ticker_worksheet(workbook: openpyxl.Workbook, ticker: str):
    for name in workbook.sheetnames:
        if name.strip().upper() == ticker.strip().upper():
            return name
    return None


DB_PERIOD_COLUMN = 1
DB_FIRST_DATA_ROW = 6


def period_to_db_format(period: str) -> str:
    quarter, year = period.split()
    return f"{year[-2:]}Q{quarter[1]}"


def find_period_row(worksheet, period: str) -> int | None:
    target = period_to_db_format(period)
    for row in worksheet.iter_rows(min_row=DB_FIRST_DATA_ROW, min_col=DB_PERIOD_COLUMN, max_col=DB_PERIOD_COLUMN):
        cell = row[0]
        value = cell.value
        if value is not None and str(value).strip().upper() == target:
            return cell.row
    return None


DB_READTHROUGH_COL = column_index_from_string("DX")
DB_VARIABLE_THEME_START_COL = DB_READTHROUGH_COL + 9
DB_VARIABLE_THEME_HEADER_ROW = 4


def format_signal_cell(signal: str, text: str, rationale: str) -> str:
    return f"{signal} {text} — {rationale}"


def get_existing_variable_themes(worksheet) -> list[str]:
    themes = []
    col = DB_VARIABLE_THEME_START_COL
    while True:
        cell = worksheet.cell(row=DB_VARIABLE_THEME_HEADER_ROW, column=col)
        if cell.value is None:
            break
        themes.append(str(cell.value).strip())
        col += 3
    return themes


def get_or_create_variable_theme_columns(worksheet, theme_name: str):
    col = DB_VARIABLE_THEME_START_COL
    while True:
        header_cell = worksheet.cell(row=DB_VARIABLE_THEME_HEADER_ROW, column=col)
        if header_cell.value is None:
            header_cell.value = theme_name
            return col, col + 1, col + 2
        if str(header_cell.value).strip() == theme_name.strip():
            return col, col + 1, col + 2
        col += 3


def write_variable_theme_row(worksheet, row: int, theme_result: dict, ranked_themes: list):
    for entry in (theme_result or {}).get("theme_analysis", []):
        theme_name = entry["theme"]
        rank_col, mgmt_col, analyst_col = get_or_create_variable_theme_columns(worksheet, theme_name)
        worksheet.cell(row=row, column=rank_col).value = (
            ranked_themes.index(theme_name) + 1 if theme_name in ranked_themes else None
        )
        worksheet.cell(row=row, column=mgmt_col).value = format_signal_cell(
            entry["mgmt"]["signal"], entry["mgmt"]["message"], entry["mgmt"]["rationale"]
        )
        worksheet.cell(row=row, column=analyst_col).value = format_signal_cell(
            entry["analyst"]["signal"], entry["analyst"]["tone"], entry["analyst"]["rationale"]
        )


def write_all_theme_results(worksheet, parsed_quarters, period_rows, themes_result, theme_delta_by_key):
    for later_item, _ in reversed(list(zip(parsed_quarters[1:], parsed_quarters[:-1]))):
        for theme_name in themes_result.get(later_item["period"], {}).get("themes", []):
            get_or_create_variable_theme_columns(worksheet, theme_name)

    for later_item, earlier_item in zip(parsed_quarters[1:], parsed_quarters[:-1]):
        delta_key = f"{later_item['label']}_vs_{earlier_item['label']}"
        ranked_themes = themes_result.get(later_item["period"], {}).get("themes", [])
        theme_result = theme_delta_by_key.get(delta_key)
        if theme_result is None:
            continue
        write_variable_theme_row(worksheet, period_rows[later_item["period"]], theme_result, ranked_themes)


# ============================================================
# UTILITIES
# ============================================================

def parse_md_metadata(filename: str):
    name = filename.rsplit(".", 1)[0]
    parts = name.split("-")
    ticker = " ".join(parts[:2]) if len(parts) >= 2 else None
    period_match = re.search(r"Q[1-4]\s+20\d{2}", name)
    period = period_match.group(0) if period_match else None
    return ticker, period


def quarter_sort_key(period: str):
    quarter, year = period.split()
    return int(year), int(quarter[1])


# ============================================================
# STREAMLIT UI
# ============================================================

st.set_page_config(page_title="Variable Theme Analysis", layout="wide")
st.title("Variable Theme Analysis")

with st.sidebar:
    api_key = st.text_input("Anthropic API key", type="password")

theme_md_files = st.file_uploader(
    "Upload telegraphic .md files (one company, 2 or more consecutive quarters)",
    type=["md"],
    accept_multiple_files=True,
)

db_file = st.file_uploader(
    "Upload DB format.xlsx (workbook to update with this analysis)",
    type=["xlsx"],
)

if theme_md_files:
    parsed_quarters = []
    for f in theme_md_files:
        ticker, period = parse_md_metadata(f.name)
        if period is None:
            st.error(f"{f.name}: could not parse a quarter/period (e.g. 'Q1 2026') from filename")
            continue
        parsed_quarters.append(
            {
                "filename": f.name,
                "ticker": ticker,
                "period": period,
                "content": f.read().decode("utf-8"),
            }
        )

    if len(parsed_quarters) < MIN_QUARTERS:
        st.warning(f"Need at least {MIN_QUARTERS} valid .md files; found {len(parsed_quarters)}.")
    else:
        parsed_quarters.sort(key=lambda item: quarter_sort_key(item["period"]))
        labels = build_quarter_labels(len(parsed_quarters))
        for label, item in zip(labels, parsed_quarters):
            item["label"] = label

        st.write({item["label"]: f"{item['ticker']} {item['period']}" for item in parsed_quarters})

        ticker = parsed_quarters[0]["ticker"]
        sheet_name = None
        period_rows = {}
        if db_file:
            workbook = openpyxl.load_workbook(db_file)
            sheet_name = find_ticker_worksheet(workbook, ticker)
            if sheet_name is None:
                st.error(f"No tab found matching ticker '{ticker}' in {db_file.name}")
            else:
                st.success(f"Found matching tab: '{sheet_name}'")
                worksheet = workbook[sheet_name]
                missing_periods = []
                for item in parsed_quarters:
                    row = find_period_row(worksheet, item["period"])
                    if row is None:
                        missing_periods.append(item["period"])
                    else:
                        period_rows[item["period"]] = row
                if missing_periods:
                    st.error(
                        f"No row found in '{sheet_name}' for: {', '.join(missing_periods)} "
                        f"(expected column A format like '{period_to_db_format(parsed_quarters[0]['period'])}')"
                    )
                else:
                    st.success(
                        "Matched rows: "
                        + ", ".join(f"{period} -> row {row}" for period, row in period_rows.items())
                    )
        else:
            st.info(f"Upload DB format.xlsx above to update the '{ticker}' tab with this analysis.")

        rows_ready = db_file and sheet_name and len(period_rows) == len(parsed_quarters)

        existing_themes = get_existing_variable_themes(worksheet) if rows_ready else []

        use_batch = st.checkbox(
            "Use batch processing (cheaper, ~50% cost)",
            help="Submits as Batches API jobs at ~50% lower API cost. "
            "Runs automatically: stage 1 identifies themes, stage 2 analyzes each delta. "
            "Keep this tab open until the download button appears.",
        )

        if use_batch:
            if st.button("Run batch", disabled=not (api_key and rows_ready)):
                client = Anthropic(api_key=api_key)
                status = st.empty()

                status.info("Submitting stage 1 batch (theme identification)...")
                batch_id = submit_stage1_batch(client, parsed_quarters, existing_themes)

                while True:
                    batch = client.messages.batches.retrieve(batch_id)
                    if batch.processing_status == "ended":
                        break
                    counts = batch.request_counts
                    status.info(
                        f"Stage 1 running — processing: {counts.processing}, "
                        f"succeeded: {counts.succeeded}, errored: {counts.errored}. Checking again in 30s..."
                    )
                    time.sleep(30)

                status.info("Fetching stage 1 results...")
                raw_results = fetch_raw_batch_results(client, batch_id)

                themes_text, themes_err, _ = raw_results.get("themes", (None, "missing", None))
                themes_result = None
                if themes_text:
                    themes_result, parse_err = parse_themes_response(themes_text)
                    if parse_err:
                        st.error(f"themes: {parse_err}")
                        with st.expander("themes raw response"):
                            st.text(themes_text)
                else:
                    st.error(f"themes batch error: {themes_err}")

                if themes_result is not None:
                    st.json(themes_result)

                    status.info("Submitting stage 2 batch (per-delta theme analysis)...")
                    stage2_id, manifest = submit_stage2_batch(client, themes_result, parsed_quarters)

                    while True:
                        batch = client.messages.batches.retrieve(stage2_id)
                        if batch.processing_status == "ended":
                            break
                        counts = batch.request_counts
                        status.info(
                            f"Stage 2 running — processing: {counts.processing}, "
                            f"succeeded: {counts.succeeded}, errored: {counts.errored}. Checking again in 30s..."
                        )
                        time.sleep(30)

                    status.info("Fetching stage 2 results and writing workbook...")
                    raw_results = fetch_raw_batch_results(client, stage2_id)

                    theme_delta_by_key = {}
                    for delta_key, ids in manifest.items():
                        theme_text, theme_err, theme_blocks = raw_results.get(ids["theme_id"], (None, "missing", None))
                        if theme_text is not None:
                            if not theme_text:
                                st.error(f"{delta_key}: batch succeeded but response contained no text (blocks: {theme_blocks})")
                            else:
                                theme_result, parse_err = parse_fenced_json_response(theme_text)
                                if parse_err:
                                    st.error(f"{delta_key}: JSON parse failed — {parse_err}")
                                    with st.expander(f"{delta_key} raw response"):
                                        st.text(theme_text)
                                else:
                                    theme_delta_by_key[delta_key] = {
                                        "result": theme_result,
                                        "ranked_themes": ids["ranked_themes"],
                                    }
                        else:
                            st.error(f"{delta_key} batch error: {theme_err}")

                    write_all_theme_results(
                        worksheet, parsed_quarters, period_rows, themes_result,
                        {k: v["result"] for k, v in theme_delta_by_key.items()},
                    )

                    st.json({k: v["result"] for k, v in theme_delta_by_key.items()})

                    output_buffer = io.BytesIO()
                    workbook.save(output_buffer)
                    status.success("Batch analysis complete.")
                    st.download_button(
                        label=f"Download updated {db_file.name}",
                        data=output_buffer.getvalue(),
                        file_name=db_file.name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="batch_dl",
                    )
                else:
                    status.error("Stage 1 failed to produce themes — cannot continue.")
            elif not api_key:
                st.info("Enter your Anthropic API key in the sidebar to begin.")

        elif st.button("Analyze", disabled=not (api_key and rows_ready)):
            client = Anthropic(api_key=api_key)

            with st.spinner("Identifying variable themes..."):
                themes_result, themes_error = analyze_themes(client, parsed_quarters, existing_themes)
            if themes_error:
                st.error(themes_error)
            else:
                st.json(themes_result)

                theme_delta_by_key = {}
                for later_item, earlier_item in zip(parsed_quarters[1:], parsed_quarters[:-1]):
                    delta_key = f"{later_item['label']}_vs_{earlier_item['label']}"
                    ranked_themes = themes_result.get(later_item["period"], {}).get("themes", [])
                    if not ranked_themes:
                        st.error(f"{delta_key}: no ranked themes found for {later_item['period']}")
                        continue
                    with st.spinner(f"Analyzing variable themes for {delta_key}..."):
                        theme_result, theme_error = analyze_theme_delta(
                            client, ranked_themes, later_item, earlier_item
                        )
                    if theme_error:
                        st.error(f"{delta_key}: {theme_error}")
                    theme_delta_by_key[delta_key] = theme_result
                    st.write(f"Variable theme analysis — {delta_key}")
                    st.json(theme_result)

                write_all_theme_results(worksheet, parsed_quarters, period_rows, themes_result, theme_delta_by_key)

                output_buffer = io.BytesIO()
                workbook.save(output_buffer)
                st.download_button(
                    label=f"Download updated {db_file.name}",
                    data=output_buffer.getvalue(),
                    file_name=db_file.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

elif not api_key:
    st.info("Enter your Anthropic API key in the sidebar to begin.")
else:
    st.info(f"Upload at least {MIN_QUARTERS} telegraphic .md files to begin.")
