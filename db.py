import json
import os
import re
from typing import Optional

import psycopg2
from openpyxl.utils import column_index_from_string, get_column_letter

# ============================================================
# CONSTANTS
# ============================================================

COMMON_THEME_SET = [
    "Top-line Growth (e.g. Revenue, Bookings, GMV, ARR, etc.)",
    "Bottom-line Expansion/Contraction (e.g. Margins, Return on Investment, Take-rate, etc.)",
    "Financial Guidance and Forward-looking Business Commentary",
]

COMMON_THEME_PREFIX = {
    COMMON_THEME_SET[0]: "topline",
    COMMON_THEME_SET[1]: "bottomline",
    COMMON_THEME_SET[2]: "guidance",
}

DB_PERIOD_COLUMN = 1
DB_DATE_COLUMN = 2
DB_FIRST_DATA_ROW = 6
DB_READTHROUGH_COL = column_index_from_string("DX")       # 128
DB_BULL_COL = DB_READTHROUGH_COL + 1
DB_BEAR_COL = DB_READTHROUGH_COL + 2
DB_COMMON_THEME_COLUMNS = {
    COMMON_THEME_SET[0]: (DB_READTHROUGH_COL + 3, DB_READTHROUGH_COL + 4),
    COMMON_THEME_SET[1]: (DB_READTHROUGH_COL + 5, DB_READTHROUGH_COL + 6),
    COMMON_THEME_SET[2]: (DB_READTHROUGH_COL + 7, DB_READTHROUGH_COL + 8),
}
DB_VARIABLE_THEME_START_COL = DB_READTHROUGH_COL + 9      # 137
DB_VARIABLE_THEME_HEADER_ROW = 4
DB_BLOOMBERG_END_COL = column_index_from_string("DW")     # 127


# ============================================================
# CONNECTION
# ============================================================

def get_connection(database_url: str = None) -> psycopg2.extensions.connection:
    return psycopg2.connect(database_url or os.environ["DATABASE_URL"])


# ============================================================
# SIGNAL PARSING
# ============================================================

def signal_to_int(signal_str: str) -> Optional[int]:
    """'↑:' / '↓:' / '→:' → 1 / -1 / 0."""
    if not signal_str:
        return None
    s = str(signal_str).strip()
    if s.startswith("↑"):
        return 1
    if s.startswith("↓"):
        return -1
    if s.startswith("→"):
        return 0
    return None


def parse_excel_signal_cell(value) -> tuple:
    """Parse '↑: text — rationale' from an Excel cell into (signal_int, text, rationale)."""
    if value is None:
        return None, None, None
    text = str(value).strip()
    match = re.match(r"^([↑↓→]):\s*(.*)", text, flags=re.DOTALL)
    if not match:
        return None, text, None
    signal = signal_to_int(match.group(1))
    rest = match.group(2).strip()
    parts = rest.split(" — ", 1)
    description = parts[0].strip() if parts else None
    rationale = parts[1].strip() if len(parts) > 1 else None
    return signal, description, rationale


# ============================================================
# PERIOD FORMAT
# ============================================================

def period_to_quarter(period: str) -> str:
    """'Q1 2026' → '26Q1'"""
    quarter, year = period.split()
    return f"{year[-2:]}Q{quarter[1]}"


# ============================================================
# EXCEL HELPERS
# ============================================================

def build_bloomberg_header_map(worksheet) -> dict:
    """Column index → display name for Bloomberg columns C (3) through DW (127).
    Uses row 4 as primary header, falls back to row 1 (Bloomberg field code)."""
    headers = {}
    for col in range(3, DB_BLOOMBERG_END_COL + 1):
        h4 = worksheet.cell(4, col).value
        h1 = worksheet.cell(1, col).value
        if h4 and str(h4).strip():
            headers[col] = str(h4).strip()
        elif h1 and str(h1).strip():
            headers[col] = str(h1).strip()
    return headers


def get_variable_themes_from_sheet(worksheet) -> list:
    """Return [(theme_name, start_col)] for variable theme columns in this sheet."""
    themes = []
    col = DB_VARIABLE_THEME_START_COL
    while True:
        v = worksheet.cell(DB_VARIABLE_THEME_HEADER_ROW, col).value
        if v is None:
            break
        themes.append((str(v).strip(), col))
        col += 3
    return themes


# ============================================================
# UPSERT FUNCTIONS
# ============================================================

def upsert_ticker(cur, ticker: str, index_name: str, currency: str, sector: str):
    cur.execute("""
        INSERT INTO tickers (ticker, index_name, currency, sector)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (ticker) DO UPDATE SET
            index_name = EXCLUDED.index_name,
            currency   = EXCLUDED.currency,
            sector     = EXCLUDED.sector
    """, (ticker, index_name, currency, sector))


def upsert_quarter(cur, ticker: str, quarter: str, date) -> int:
    date_val = date.date() if hasattr(date, "date") else date
    cur.execute("""
        INSERT INTO quarters (ticker, quarter, date)
        VALUES (%s, %s, %s)
        ON CONFLICT (ticker, quarter) DO UPDATE SET date = EXCLUDED.date
        RETURNING id
    """, (ticker, quarter, date_val))
    return cur.fetchone()[0]


def get_quarter_id(cur, ticker: str, quarter: str) -> Optional[int]:
    cur.execute(
        "SELECT id FROM quarters WHERE ticker = %s AND quarter = %s",
        (ticker, quarter),
    )
    row = cur.fetchone()
    return row[0] if row else None


def upsert_quarter_signals(cur, quarter_id: int, data: dict):
    cur.execute("""
        INSERT INTO quarter_signals (
            quarter_id,
            read_through_signal, read_through_description,
            bull_signal, bull_description,
            bear_signal, bear_description,
            topline_mgmt_signal, topline_mgmt_description,
            topline_analyst_signal, topline_analyst_description,
            bottomline_mgmt_signal, bottomline_mgmt_description,
            bottomline_analyst_signal, bottomline_analyst_description,
            guidance_mgmt_signal, guidance_mgmt_description,
            guidance_analyst_signal, guidance_analyst_description
        ) VALUES (
            %(quarter_id)s,
            %(read_through_signal)s, %(read_through_description)s,
            %(bull_signal)s, %(bull_description)s,
            %(bear_signal)s, %(bear_description)s,
            %(topline_mgmt_signal)s, %(topline_mgmt_description)s,
            %(topline_analyst_signal)s, %(topline_analyst_description)s,
            %(bottomline_mgmt_signal)s, %(bottomline_mgmt_description)s,
            %(bottomline_analyst_signal)s, %(bottomline_analyst_description)s,
            %(guidance_mgmt_signal)s, %(guidance_mgmt_description)s,
            %(guidance_analyst_signal)s, %(guidance_analyst_description)s
        )
        ON CONFLICT (quarter_id) DO UPDATE SET
            read_through_signal            = EXCLUDED.read_through_signal,
            read_through_description       = EXCLUDED.read_through_description,
            bull_signal                    = EXCLUDED.bull_signal,
            bull_description               = EXCLUDED.bull_description,
            bear_signal                    = EXCLUDED.bear_signal,
            bear_description               = EXCLUDED.bear_description,
            topline_mgmt_signal            = EXCLUDED.topline_mgmt_signal,
            topline_mgmt_description       = EXCLUDED.topline_mgmt_description,
            topline_analyst_signal         = EXCLUDED.topline_analyst_signal,
            topline_analyst_description    = EXCLUDED.topline_analyst_description,
            bottomline_mgmt_signal         = EXCLUDED.bottomline_mgmt_signal,
            bottomline_mgmt_description    = EXCLUDED.bottomline_mgmt_description,
            bottomline_analyst_signal      = EXCLUDED.bottomline_analyst_signal,
            bottomline_analyst_description = EXCLUDED.bottomline_analyst_description,
            guidance_mgmt_signal           = EXCLUDED.guidance_mgmt_signal,
            guidance_mgmt_description      = EXCLUDED.guidance_mgmt_description,
            guidance_analyst_signal        = EXCLUDED.guidance_analyst_signal,
            guidance_analyst_description   = EXCLUDED.guidance_analyst_description
    """, {"quarter_id": quarter_id, **data})


def _combine_description(text: Optional[str], rationale: Optional[str]) -> Optional[str]:
    if text and rationale:
        return f"{text} — {rationale}"
    return text or rationale


def upsert_variable_theme(
    cur,
    quarter_id: int,
    theme_name: str,
    rank: Optional[int],
    mgmt_signal: Optional[int],
    mgmt_description: Optional[str],
    analyst_signal: Optional[int],
    analyst_description: Optional[str],
):
    cur.execute("""
        INSERT INTO variable_themes (
            quarter_id, theme_name, rank,
            mgmt_signal, mgmt_description,
            analyst_signal, analyst_description
        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (quarter_id, theme_name) DO UPDATE SET
            rank                = EXCLUDED.rank,
            mgmt_signal         = EXCLUDED.mgmt_signal,
            mgmt_description    = EXCLUDED.mgmt_description,
            analyst_signal      = EXCLUDED.analyst_signal,
            analyst_description = EXCLUDED.analyst_description
    """, (
        quarter_id, theme_name, rank,
        mgmt_signal, mgmt_description,
        analyst_signal, analyst_description,
    ))


def upsert_external_data(cur, ticker: str, quarter: str, data: dict):
    cur.execute("""
        INSERT INTO external_data (ticker, quarter, data)
        VALUES (%s, %s, %s::jsonb)
        ON CONFLICT (ticker, quarter) DO UPDATE SET
            data = external_data.data || EXCLUDED.data
    """, (ticker, quarter, json.dumps(data)))


def get_existing_variable_themes(cur, ticker: str) -> list:
    cur.execute("""
        SELECT DISTINCT vt.theme_name
        FROM variable_themes vt
        JOIN quarters q ON q.id = vt.quarter_id
        WHERE q.ticker = %s
        ORDER BY vt.theme_name
    """, (ticker,))
    return [row[0] for row in cur.fetchall()]


# ============================================================
# EXCEL → SUPABASE: STRUCTURE SYNC
# ============================================================

def sync_workbook_structure(conn, workbook) -> dict:
    """Upsert tickers, quarters, and Bloomberg external_data from the uploaded workbook.

    Returns {ticker: {quarter_str: quarter_id}} for all sheets processed.
    """
    cur = conn.cursor()
    quarter_id_map = {}

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ticker = str(ws.cell(1, 2).value or "").strip()
        if not ticker:
            continue

        index_name = str(ws.cell(1, 1).value or "").strip() or None
        currency   = str(ws.cell(2, 1).value or "").strip() or None
        sector     = str(ws.cell(2, 2).value or "").strip() or None

        upsert_ticker(cur, ticker, index_name, currency, sector)

        bloomberg_headers = build_bloomberg_header_map(ws)
        quarter_id_map[ticker] = {}

        for row in range(DB_FIRST_DATA_ROW, ws.max_row + 1):
            quarter_val = ws.cell(row, DB_PERIOD_COLUMN).value
            if quarter_val is None:
                break
            quarter = str(quarter_val).strip()
            date_val = ws.cell(row, DB_DATE_COLUMN).value
            qid = upsert_quarter(cur, ticker, quarter, date_val)
            quarter_id_map[ticker][quarter] = qid

            bloomberg_data = {}
            for col, header in bloomberg_headers.items():
                val = ws.cell(row, col).value
                if val is None:
                    continue
                str_val = str(val).strip()
                if str_val in ("", ", "):
                    continue
                bloomberg_data[header] = val.isoformat() if hasattr(val, "isoformat") else val
            if bloomberg_data:
                upsert_external_data(cur, ticker, quarter, bloomberg_data)

    conn.commit()
    cur.close()
    return quarter_id_map


# ============================================================
# ANALYSIS RESULTS → SUPABASE
# ============================================================

def _blank_signals() -> dict:
    """Return a dict with all quarter_signals fields set to None."""
    fields = [
        "read_through_signal", "read_through_description",
        "bull_signal", "bull_description",
        "bear_signal", "bear_description",
    ]
    for prefix in ("topline", "bottomline", "guidance"):
        fields += [
            f"{prefix}_mgmt_signal", f"{prefix}_mgmt_description",
            f"{prefix}_analyst_signal", f"{prefix}_analyst_description",
        ]
    return {f: None for f in fields}


def write_analysis_results(
    conn,
    ticker: str,
    quarter_id_map: dict,
    parsed_quarters: list,
    readthrough_result: Optional[dict],
    bullbear_result: Optional[dict],
    theme_delta_by_key: dict,
):
    """Write theme analysis results to quarter_signals and variable_themes."""
    cur = conn.cursor()

    rt_lookup = {e["period"]: e for e in (readthrough_result or {}).get("historical_analysis", [])}
    bb_lookup = {e["period"]: e for e in (bullbear_result or {}).get("historical_analysis", [])}

    for later_item, earlier_item in zip(parsed_quarters[1:], parsed_quarters[:-1]):
        delta_key = f"{later_item['label']}_vs_{earlier_item['label']}"
        later_quarter = period_to_quarter(later_item["period"])
        quarter_id = quarter_id_map.get(later_quarter)
        if quarter_id is None:
            continue

        rt  = rt_lookup.get(delta_key)
        bb  = bb_lookup.get(delta_key)
        entry = theme_delta_by_key.get(delta_key)

        signals = _blank_signals()
        signals["quarter_id"] = quarter_id

        if rt:
            signals["read_through_signal"]      = signal_to_int(rt.get("signal"))
            signals["read_through_description"] = _combine_description(rt.get("read_through"), rt.get("rationale"))

        if bb:
            signals["bull_signal"]      = signal_to_int(bb["bull"].get("topic"))
            signals["bull_description"] = _combine_description(bb["bull"].get("expectation"), bb["bull"].get("context"))
            signals["bear_signal"]      = signal_to_int(bb["bear"].get("topic"))
            signals["bear_description"] = _combine_description(bb["bear"].get("expectation"), bb["bear"].get("context"))

        if entry:
            common_entries = {e["theme"]: e for e in entry["common"].get("theme_analysis", [])}
            for theme_name, prefix in COMMON_THEME_PREFIX.items():
                e = common_entries.get(theme_name)
                if not e:
                    continue
                signals[f"{prefix}_mgmt_signal"]          = signal_to_int(e["mgmt"].get("signal"))
                signals[f"{prefix}_mgmt_description"]     = _combine_description(e["mgmt"].get("message"), e["mgmt"].get("rationale"))
                signals[f"{prefix}_analyst_signal"]       = signal_to_int(e["analyst"].get("signal"))
                signals[f"{prefix}_analyst_description"]  = _combine_description(e["analyst"].get("tone"), e["analyst"].get("rationale"))

        upsert_quarter_signals(cur, quarter_id, signals)

        if entry:
            ranked_themes = entry.get("ranked_themes", [])
            for vt in entry["ranked"].get("theme_analysis", []):
                theme_name = vt["theme"]
                rank = (ranked_themes.index(theme_name) + 1) if theme_name in ranked_themes else None
                upsert_variable_theme(
                    cur, quarter_id, theme_name, rank,
                    signal_to_int(vt["mgmt"].get("signal")),
                    _combine_description(vt["mgmt"].get("message"), vt["mgmt"].get("rationale")),
                    signal_to_int(vt["analyst"].get("signal")),
                    _combine_description(vt["analyst"].get("tone"), vt["analyst"].get("rationale")),
                )

    conn.commit()
    cur.close()
